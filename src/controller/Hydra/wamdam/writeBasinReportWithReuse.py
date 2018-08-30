
import xlsxwriter
import xlrd as excel
from openpyxl import load_workbook

full_path = 'SumFinal_BasinReportWithReuse_result.xlsx'
workbook1 = xlsxwriter.Workbook(full_path)

cell_format_field_name = workbook1.add_format()
cell_format_field_name.set_font_size(14)
cell_format_field_name.set_bold()

workbook1.add_worksheet('SumFinal_BasinReportWithReuse')
workbook1.close()
##########################################################

book = excel.open_workbook('SumFinal_BasinReportWithReuse.xlsx')
try:
    input_sheet = book.sheet_by_name('SumFinal_BasinReportWithReuse')
except:
    raise Exception('Input sheet {} not found in the provided Excel file \n Please select valid excel file'.format('SumFinal_BasinReportWithReuse'))

# get total cells of input sheet.
rows = [rows for rows in input_sheet.get_rows()]

# get headers
headers = rows[1][9:]

# get total value
data = [row[0:] for row in rows[2:]]

data_to_write = []

# open sheet to write new xlsx
workbook = load_workbook(full_path)
sheet = workbook.get_sheet_by_name('SumFinal_BasinReportWithReuse')

# ready data to write
# firstly, loop headers
for i in range(len(headers)):
    for j, d in enumerate(data):
        data_to_write_temp = [str(int(d[0].value)), d[1].value, headers[i].value, str(d[i + 9].value)]
        data_to_write.append(data_to_write_temp)
#write
for rowID, row in enumerate(data_to_write):
    for colID, cell in enumerate(row):
        try:
            sheet.cell(row=rowID + 3, column=colID + 1, value=unicode(cell))
        except Exception as e:
            raise Exception(e)

workbook.save(full_path)