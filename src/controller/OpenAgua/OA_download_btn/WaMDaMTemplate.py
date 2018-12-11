# maybe write the script here? then send it to the viewer dlg_ExportScenarioDataToHydra  ?


import xlsxwriter, os, Global
from openpyxl import load_workbook

import pandas as pd

class ExportTemplate (object):
	def __init__(self, full_path):

		##############################################
		# create xlsx file with "full_path".
		workbook1 = xlsxwriter.Workbook(full_path)
		workbook1.add_worksheet('1.1_Organiz&People')
		workbook1.add_worksheet('1.2_Sources&Methods')
		workbook1.add_worksheet('2.1_ResourceTypes&ObjectTypes')
		workbook1.add_worksheet('2.2_Attributes')
		workbook1.add_worksheet('3.1_Networks&Scenarios')
		workbook1.add_worksheet('3.2_Nodes')
		workbook1.add_worksheet('3.3_Links')
		workbook1.add_worksheet('4_NumericValues')
		workbook1.add_worksheet('4_FreeText')
		workbook1.add_worksheet('4_CategoricalValues')
		workbook1.add_worksheet('4_SeasonalNumericValues')
		workbook1.add_worksheet('4_TimeSeries')
		workbook1.add_worksheet('4_TimeSeriesValues')
		workbook1.add_worksheet('4_MultiAttributeSeries')
		workbook1.add_worksheet('ObjectCategory')
		workbook1.add_worksheet('AttributeCategory')
		workbook1.add_worksheet('InstanceCategory')
		workbook1.close()
		##########################################################
		
		self.workbook = load_workbook(full_path)
		self.file_path = full_path

		self.pandas_excel_writer = pd.ExcelWriter(full_path, engine='xlsxwriter')

	def exportOrganizations(self, organization_list):
		organization_list.to_excel(self.pandas_excel_writer, sheet_name="1.1_Organiz&People", startrow=2, startcol=0, header=True,
					index=False)

	def exportPeople(self, people_list):
		df = pd.DataFrame(people_list)
		df.to_excel(self.pandas_excel_writer, sheet_name="1.1_Organiz&People", startrow=9, startcol=0, header=True,
					index=False)

	def exportSources(self, source_list):
		source_list.to_excel(self.pandas_excel_writer, sheet_name="1.2_Sources&Methods", startrow=8, startcol=0, header=True,
					index=False)

	def exportMethods(self, method_list):
		df = pd.DataFrame(method_list)
		df.to_excel(self.pandas_excel_writer, sheet_name="1.2_Sources&Methods", startrow=19, startcol=0, header=True,
					index=False)

	def exportMasterNetwork(self, network_data_result):

		network_data_result.to_excel(self.pandas_excel_writer, sheet_name="3.1_Networks&Scenarios", startrow=8, startcol=0, header=True,
					index=False)


	def exportScenario(self, scenarios_data_result):
		scenarios_data_result.to_excel(self.pandas_excel_writer, sheet_name="3.1_Networks&Scenarios", startrow=18, startcol=0,
					header=True,
					index=False)

	def exportNodes(self, nodes_data_result):
		nodes_data_result.to_excel(self.pandas_excel_writer, sheet_name="3.2_Nodes", startrow=8, startcol=0, header=True, index=False)

	def exportLinkes(self, links_frame_result):
		links_frame_result.to_excel(self.pandas_excel_writer, sheet_name="3.3_Links", startrow=8, startcol=0, header=True,
					index=False)

	def exportObjecttypes(self, data_result):
		df = pd.DataFrame(data_result)
		df.to_excel(self.pandas_excel_writer, sheet_name="2.1_ResourceTypes&ObjectTypes", header=True,
					index=False)

	def exportResourcesType(self, resources_result , objects_result):
		if not resources_result.empty:
			resources_result.to_excel(self.pandas_excel_writer, sheet_name="2.1_ResourceTypes&ObjectTypes", startrow=8, startcol=0,
					header=True,
					index=False)
		if not objects_result.empty:
			objects_result.to_excel(self.pandas_excel_writer, sheet_name="2.1_ResourceTypes&ObjectTypes", startrow=16, startcol=0,
					header=True,
					index=False)

	def exportAttributes(self, attributes_result):
		attributes_result.to_excel(self.pandas_excel_writer, sheet_name="2.2_Attributes", startrow=9, startcol=0,
					header=True,
					index=False)


	def exportFreeText(self, data_list):
		df = data_list
		df.to_excel(self.pandas_excel_writer, sheet_name="4_FreeText", startrow=8, startcol=0,
					header=True,
					index=False)

	def exportNumericValue(self, data_list):
		df = data_list
		df.to_excel(self.pandas_excel_writer, sheet_name="4_NumericValues", startrow=8, startcol=0,
					header=True,
					index=False)

	def exportElectronicFiles(self, data_list):
		df = pd.DataFrame(data_list)
		df.to_excel(self.pandas_excel_writer, sheet_name="4_ElectronicFiles", startrow=8, startcol=0,
					header=True,
					index=False)

	def exportSeasonal(self, data_list):


		data_list.to_excel(self.pandas_excel_writer, sheet_name="4_SeasonalNumericValues", startrow=8, startcol=0,
					header=True,
					index=False)


	def exportTimeSeries(self, data_list):
		data_list.to_excel(self.pandas_excel_writer, sheet_name="4_TimeSeries", startrow=8, startcol=0,
					header=True,
					index=False)


	def exportTimeSeriesValues(self, data_list):
		df = data_list
		df.to_excel(self.pandas_excel_writer, sheet_name="4_TimeSeriesValues", startrow=8, startcol=0,
					header=True,
					index=False)


	def exportCategoricalValues(self, data_list):
		df = pd.DataFrame(data_list)
		df.to_excel(self.pandas_excel_writer, sheet_name="4_CategoricalValues", startrow=8, startcol=0,
					header=True,
					index=False)

	def exportMulti(self, up_table_column_result, bottom_table_result):
		up_table_column_result.to_excel(self.pandas_excel_writer, sheet_name="4_MultiAttributeSeries", startrow=3, startcol=5,
					header=True,
					index=False)

		bottom_table_result.to_excel(self.pandas_excel_writer, sheet_name="4_MultiAttributeSeries", startrow=17, startcol=0,
					header=True,
					index=False)


	def write2excel(self, resutl_list, startRow, sheet_name):
		try:
			for row_id, row in enumerate(resutl_list):
				for col_id, cell in enumerate(row):
					sheet.cell(row=row_id + startRow, column=col_id + 1, value=cell)

				print "writing row id: " + str(row_id)

				self.workbook.save(self.file_path)

		except Exception as e:
			print e
			raise Exception(e.message)