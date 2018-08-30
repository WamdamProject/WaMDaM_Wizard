# maybe write the script here? then send it to the viewer dlg_ExportScenarioDataToHydra  ?

import xlsxwriter, os, Global, csv
from openpyxl import load_workbook
from collections import OrderedDict

class SaveToExcelFromCSV (object):
	def __init__(self, full_dir, dir_name):

		# create class variables to write data in xlsx file.
		# self.setup = DB_Setup()
		# self.session = self.setup.get_session()
        #
		# self.dataStructure = GetDataStructure()
		# self.getDataValues = GetDataValues()
		##############################################

		# create xlsx file with "full_path".
		workbook1 = xlsxwriter.Workbook("{}/{}.xlsx".format(full_dir, dir_name))

		cell_format_field_name = workbook1.add_format()
		cell_format_field_name.set_font_size(14)
		cell_format_field_name.set_bold()

		cell_format_values = workbook1.add_format()
		cell_format_values.set_border()
		cell_format_values.set_bg_color('#FFFFCC')

		worksheet = workbook1.add_worksheet('1.1_Organiz&People')
		fieldNames = [["OrganizationName", "OrganizationType", "OrganizationWebpage", "Description"]]
		Global.SetValuesAndFormatCells(worksheet, fieldNames, 0, 0, 0, 0, cell_format_field_name)
		Global.SetValuesAndFormatCells(worksheet, [], 1, 5, 0, 4, cell_format_values)
		fieldNames = [["PersonName", "Address", "Email", "Phone", "PersonWebpage", "Position", "OrganizationName"]]
		Global.SetValuesAndFormatCells(worksheet, fieldNames, 7, 0, 0, 0, cell_format_field_name)
		Global.SetValuesAndFormatCells(worksheet, [], 8, 10, 0, 7, cell_format_values)

		workbook1.add_worksheet('1.2_Sources&Methods')
		workbook1.add_worksheet('2.1_ResourceTypes&ObjectTypes')
		workbook1.add_worksheet('2.2_Attributes')
		workbook1.add_worksheet('3.1_Networks&Scenarios')
		workbook1.add_worksheet('3.2_Nodes')
		workbook1.add_worksheet('3.3_Links')
		workbook1.add_worksheet('4_NumericValues')
		workbook1.add_worksheet('4_CategoricalValues')
		workbook1.add_worksheet('4_SeasonalNumericValues')
		workbook1.add_worksheet('4_TimeSeries')
		workbook1.add_worksheet('4_TimeSeriesValues')
		workbook1.add_worksheet('4_MultiAttributeSeries')
		workbook1.add_worksheet('4_FreeText')
		workbook1.add_worksheet('ObjectCategory')
		workbook1.add_worksheet('AttributeCategory')
		workbook1.add_worksheet('InstanceCategory')
		workbook1.close()
		##########################################################
		
		self.workbook = load_workbook("{}/{}.xlsx".format(full_dir, dir_name))
		self.excel_file_path = "{}/{}.xlsx".format(full_dir, dir_name)
		self.file_dir = full_dir

	# def exportOrganizations(self, selectedResourceType):
	# 	organization_list = self.dataStructure.GetOrganizations(selectedResourceType)
	# 	self.write2excel(organization_list, 2, '1.1_Organiz&People')
    #
	# def exportPeople(self, selectedResourceType):
	# 	people_list = self.dataStructure.GetPeople(selectedResourceType)
	# 	self.write2excel(people_list, 9, '1.1_Organiz&People')
    #
	# def exportSources(self, selectedResourceType):
	# 	source_list = self.dataStructure.GetSources(selectedResourceType)
	# 	self.write2excel(source_list, 8, '1.2_Sources&Methods')
    #
	# def exportMethods(self, selectedResourceType):
	# 	method_list = self.dataStructure.GetMethods(selectedResourceType)
	# 	self.write2excel(method_list, 19, '1.2_Sources&Methods')
    #
	# def exportMasterNetwork(self, selectedResourceType):
	# 	names, network_data_result = self.dataStructure.getMasterNetwork(selectedResourceType)
	# 	self.write2excel(network_data_result, 10, '3.1_Networks&Scenarios')
    #
	# def exportScenario(self, selectedResourceType, selectedMasterNetworkName):
	# 	names, scenarios_data_result = self.dataStructure.getScenario(selectedResourceType, selectedMasterNetworkName)
	# 	self.write2excel(scenarios_data_result, 20, '3.1_Networks&Scenarios')

	def exportNodes(self):
		nodes_data_result = []
		full_path = "{}/{}".format(self.file_dir, "nodes.csv")
		f = open(full_path)
		csv_items = csv.DictReader(f)
		for i, row in enumerate(csv_items):
			row_data = ['', '', '', '', '', '', '', '', '', '']
			row_data.insert(0, row[" Type"])
			row_data.insert(1, row["Name"])
			row_data.insert(7, row[" x"])
			row_data.insert(8, row[" y"])
			nodes_data_result.append(row_data)
		self.write2excel(nodes_data_result, 2, '3.2_Nodes')

	def exportLinkes(self):
		nodes_data_result = []
		full_path = "{}/{}".format(self.file_dir, "links.csv")
		f = open(full_path)
		csv_items = csv.DictReader(f)
		for i, row in enumerate(csv_items):
			row_data = ['', '', '', '', '', '', '', '', '', '']
			row_data.insert(0, row[" Type"])
			row_data.insert(1, row["Name"])
			row_data.insert(6, row[" from"])
			row_data.insert(7, row[" to"])
			nodes_data_result.append(row_data)
		self.write2excel(nodes_data_result, 2, '3.3_Links')

	# def exportLinkes(self, selectedResourceType, selectedMasterNetworkName, selectedScenarioName):
	# 	links_data_result = self.dataStructure.getLinkes(selectedResourceType, selectedMasterNetworkName,selectedScenarioName)
	# 	self.write2excel(links_data_result, 11, '3.3_Links')
    #
	# def exportObjecttypes(self, selectedResourceType):
	# 	data_result = self.dataStructure.getObjecttypes(selectedResourceType)
	# 	self.write2excel(data_result, 18, '2.1_ResourceTypes&ObjectTypes')
    #
	# def exportDatasetType(self, selectedResourceType):
	# 	dataset_result = self.dataStructure.getResourceType(selectedResourceType)
	# 	self.write2excel(dataset_result, 10, '2.1_ResourceTypes&ObjectTypes')
    #
	# def exportAttributes(self, selectedResourceType):
	# 	attributes_result = self.dataStructure.getAttributes(selectedResourceType)
	# 	self.write2excel(attributes_result, 11, '2.2_Attributes')
    #
	# # def exportDualValues(self, selectedResourceType, selectedMasterNetworkName, selectedScenarioName):
	# # 	self.getDataValues.exportDualValuesSheet1(selectedResourceType, selectedMasterNetworkName, selectedScenarioName, self.file_path)
    #
	# def exportTextConrolled(self, selectedResourceType, selectedMasterNetworkName, selectedScenarioName):
	# 	self.getDataValues.exportTextConrolledSheet1(selectedResourceType, selectedMasterNetworkName, selectedScenarioName, self.file_path)
    #
	# def exportNumericValue(self, selectedResourceType, selectedMasterNetworkName, selectedScenarioName):
	# 	self.getDataValues.exportNumericValuesheet1(selectedResourceType, selectedMasterNetworkName, selectedScenarioName, self.file_path)
    #
	# def exportElectronicFiles(self, selectedResourceType, selectedMasterNetworkName, selectedScenarioName):
	# 	self.getDataValues.exportElectronicFilesSheet1(selectedResourceType, selectedMasterNetworkName, selectedScenarioName, self.file_path)
    #
	# def exportSeasonal(self, selectedResourceType, selectedMasterNetworkName, selectedScenarioName):
	# 	self.getDataValues.exportSeasonalSheet1(selectedResourceType, selectedMasterNetworkName, selectedScenarioName, self.file_path)
    #
	# def exportTimeSeries(self, selectedResourceType, selectedMasterNetworkName, selectedScenarioName):
	# 	self.getDataValues.exportTimeSeriesSheet1(selectedResourceType, selectedMasterNetworkName, selectedScenarioName, self.file_path)
    #
	# def exportMulti(self, selectedResourceType, selectedMasterNetworkName, selectedScenarioName):
	# 	self.getDataValues.exportMultiSheet1(selectedResourceType, selectedMasterNetworkName, selectedScenarioName, self.file_path)

	def write2excel(self, resutl_list, startRow, sheet_name):
		try:
			try:
				sheet = self.workbook.get_sheet_by_name(sheet_name)
			except:
				raise Exception("Please select a valid Excel File")
			for row_id, row in enumerate(resutl_list):
				for col_id, cell in enumerate(row):
					sheet.cell(row=row_id + startRow, column=col_id + 1, value=cell)

				self.workbook.save(self.excel_file_path)

		except Exception as e:
			print e
			raise Exception(e.message)


if __name__ == '__main__':
	saveToExcelFromCSV = SaveToExcelFromCSV("{}/{}".format(os.path.expanduser("~/Desktop"), 'network_St._Croix/Baseline'), 'Baseline')
	saveToExcelFromCSV.exportNodes()
	saveToExcelFromCSV.exportLinkes()