
from ..ConnectDB_ParseExcel import DB_Setup
from ..ConnectDB_ParseExcel import SqlAlchemy as sq
from openpyxl import load_workbook
from xlrd import open_workbook
# from xlutils.copy import copy

'''
This class is used to get data values.
'''
class GetDataValues(object):
    def __init__(self):
        self.setup = DB_Setup()
        self.session = self.setup.get_session()
    def GetCV_ObjectType(self):
        '''
        This method is used to get all names of CV_ObjectType.
        :return: list of CV_ObjectType's name
        '''
        result = self.session.query(sq.CV_ObjectType.Name).all()
        # complete_result = list()
        nameResult = list()
        for row in result:
            isExisting = False
            for name in nameResult:
                if name == row.Name:
                    isExisting = True
                    break
            if not isExisting:
                nameResult.append(row.Name)
        return nameResult

    def getAttribute(self):
        '''
        This method is used to get all names of CV_AttributeName.
        :return: list of CV_AttributeName's name
        '''
        result = self.session.query(sq.CV_AttributeName.Name).all()
        # complete_result = list()
        nameResult = list()
        for row in result:
            isExisting = False
            for name in nameResult:
                if name == row.Name:
                    isExisting = True
                    break
            if not isExisting:
                nameResult.append(row.Name)
        return nameResult

    def getCVinstanceName(self):
        '''
        This method is used to get all names of CV_InstanceName.
        :return: list of CV_InstanceName's name
        '''
        result = self.session.query(sq.CV_InstanceName.Name).all()
        # complete_result = list()
        nameResult = list()
        for row in result:
            isExisting = False
            for name in nameResult:
                if name == row.Name:
                    isExisting = True
                    break
            if not isExisting:
                nameResult.append(row.Name)
        return nameResult
    def check_validation(self):
        '''
        This method is used to check validations of bellow tables within db.
        :return: None
        '''
        table_list = ["CategoricalValues", "NumericValues", "CV_ElectronicFormat", "SeasonalNumericValues", "TimeSeries",
                      "MultiAttributeSeriesValues"]
        for table_name in table_list:
            recordCountResult = self.session.execute('SELECT COUNT(*) FROM {};'.format(table_name))
            i = 0
            for n in recordCountResult:
                i = int(n[0])
            if i == 0:
                raise Exception('Warning!\n{} table is empty. Please fill this table.'.format(table_name))

    # def exportDualValuesSheet(self, selectedType, selectedAttribute, selectedInstance, excelPath):
    #     '''
    #     This method is used to get data making DualValues_table.
    #     :param selectedType: selected Object Type
    #     :param selectedAttribute: controlled Attribute
    #     :param selectedInstance: controlled Instance Name
    #     :param excelPath: full path of excel file to export data
    #     :return: None
    #     '''
    #     try:
    #         # result = self.session.query(sq.ResourceTypes.ResourceTypeAcronym, sq.Attributes.AttributeName, sq.Attributes.UnitNameCV,
    #         #                 sq.Instances.InstanceName, sq.MasterNetworks.MasterNetworkName,
    #         #                 sq.Scenarios.ScenarioName, sq.Sources.SourceName, sq.Methods.MethodName,
    #         #                 sq.DualValues.DualValuesValueMeaningCV)\
    #         #     .join(sq.ObjectTypes, sq.ObjectTypes.ResourceTypeID == sq.ResourceTypes.ResourceTypeID)\
    #         #     .join(sq.Attributes, sq.Attributes.ObjectTypeID == sq.ObjectTypes.ObjectTypeID)\
    #         #     .join(sq.Mapping, sq.Mapping.AttributeID == sq.Attributes.AttributeID)\
    #         #     .join(sq.ValuesMapper, sq.ValuesMapper.ValuesMapperID == sq.Mapping.ValuesMapperID)\
    #         #     .join(sq.ScenarioMapping, sq.ScenarioMapping.MappingID == sq.Mapping.MappingID)\
    #         #     .join(sq.Scenarios, sq.Scenarios.ScenarioID == sq.ScenarioMapping.ScenarioID)\
    #         #     .join(sq.MasterNetworks, sq.MasterNetworks.MasterNetworkID == sq.Scenarios.MasterNetworkID)\
    #         #     .join(sq.Methods, sq.Methods.MethodID == sq.Mapping.MethodID)\
    #         #     .join(sq.Sources, sq.Sources.SourceID == sq.Mapping.SourceID)\
    #         #     .join(sq.Instances, sq.Instances.InstanceID == sq.Mapping.InstanceID)\
    #         #     .join(sq.DualValues, sq.DualValues.ValuesMapperID == sq.ValuesMapper.ValuesMapperID)\
    #         #     .join(sq.CV_DualValueMeaning, sq.CV_DualValueMeaning.Name == sq.DualValues.DualValuesValueMeaningCV)\
    #         #     .filter(sq.Attributes.AttributeDataTypeCV=='DualValues')\
    #         #     .all()
    #
    #         sql = 'SELECT AttributeName, SourceName, InstanceName,MasterNetworkName,ScenarioName,MethodName,dualvaluemeaningCV ' \
    #             'FROM "ResourceTypes" '\
    #             'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
    #             'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
    #             'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
    #             'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
    #             'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
    #             'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
    #             'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
    #             'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
    #             'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
    #             'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
    #             'LEFT JOIN "DualValues" ON "DualValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
    #             'LEFT JOIN "CV_DualValueMeaning" ON "CV_DualValueMeaning"."Name"= "DualValues"."dualvaluemeaningCV" '\
    #             'WHERE "AttributeDataTypeCV"="DualValues" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
    #             .format( selectedType, selectedInstance, selectedAttribute)
    #
    #         result = self.session.execute(sql)
    #         # nameResult = list()
    #         complete_result = list()
    #         for row in result:
    #             # isExisting = False
    #             # for name in nameResult:
    #             #     if name == row.InstanceName:
    #             #         isExisting = True
    #             #         break
    #             # if not isExisting:
    #             #     nameResult.append(row.InstanceName)
    #             complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
    #                                     row.AttributeName, row.SourceName, row.MethodName,
    #                                     row.dualvaluemeaningCV])
    #         self.isMatching_query(complete_result, "DualValues")
    #         self.write2excel(complete_result, '4_DualValues', 9, 10, excelPath)
    #     except Exception as  e:
    #         print e
    #         raise Exception('Error occurred in reading Data Structure.\n' + e.message)

    def exportCategoricalValuesSheet(self, selectedType='', selectedAttribute='', selectedInstance='', excelPath=''):
        '''
        This method is used to get data making CategoricalValues_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedType == '' and selectedAttribute == '' and selectedInstance == '':
                sql = 'SELECT Attributes.AttributeName, ObjectType, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName,CategoricalValueCV ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "CategoricalValues" ON "CategoricalValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'LEFT JOIN "CV_Categorical" ON "CV_Categorical"."Name"= "CategoricalValues"."CategoricalValueCV" '\
                    'WHERE "AttributeDataTypeCV"="CategoricalValues"'
            else:
                sql = 'SELECT Attributes.AttributeName, ObjectType, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName,CategoricalValueCV ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "CategoricalValues" ON "CategoricalValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'LEFT JOIN "CV_Categorical" ON "CV_Categorical"."Name"= "CategoricalValues"."CategoricalValueCV" '\
                    'WHERE "AttributeDataTypeCV"="CategoricalValues" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                    .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.CategoricalValueCV])
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "CategoricalValues")
                    self.write2excel(complete_result, '4_CategoricalValues', 13, 10, excelPath)

            return complete_result
        except Exception as  e:
            print e
            raise Exception('Error occured in reading Data Structure.\n' + e.message)

    def exportNumericValuesheet(self, selectedType='', selectedAttribute='', selectedInstance='', excelPath=''):
        '''
        This method is used to get data making NumericValues_table.
        :param selectedType: selected Object Type
        :param selecttedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedType == '' and selectedAttribute == '' and selectedInstance == '':
                sql = 'SELECT "ResourceTypes"."ResourceType", ObjectType,AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName, NumericValue ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "NumericValues" ON "NumericValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE "AttributeDataTypeCV"="Parameter"'
            else:
                sql = 'SELECT "ResourceTypes"."ResourceType", ObjectType,AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName, NumericValue ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "NumericValues" ON "NumericValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE "AttributeDataTypeCV"="Parameter" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                    .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.NumericValue])
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "Parameter")
                    self.write2excel(complete_result, '4_Parameter', 10, 10, excelPath)

            return complete_result
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    # def exportElectronicFilesSheet(self, selectedType='', selectedAttribute='', selectedInstance='', excelPath=''):
    #     '''
    #     This method is used to get data making ElectronicFiles_table.
    #     :param selectedType: selected Object Type
    #     :param selectedAttribute: controlled Attribute
    #     :param selectedInstance: controlled Instance Name
    #     :param excelPath: full path of excel file to export data
    #     :return: None
    #     '''
    #     try:
    #         if selectedType == '' and selectedAttribute == '' and selectedInstance == '':
    #             sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
    #                   'ScenarioName,MethodName,FileName, ElectronicFileFormatCV, "File"."Description" ' \
    #                 'FROM "ResourceTypes" '\
    #                 'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
    #                 'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
    #                 'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
    #                 'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
    #                 'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
    #                 'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
    #                 'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
    #                 'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
    #                 'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
    #                 'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
    #                 'LEFT JOIN "File" ON "File"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
    #                 'Left JOIN "CV_ElectronicFormat" ON "CV_ElectronicFormat"."Name"="File"."ElectronicFileFormatCV" '\
    #                 'WHERE "AttributeDataTypeCV"="File"'
    #         else:
    #             sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
    #                   'ScenarioName,MethodName,FileName, ElectronicFileFormatCV, "File"."Description" ' \
    #                 'FROM "ResourceTypes" '\
    #                 'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
    #                 'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
    #                 'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
    #                 'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
    #                 'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
    #                 'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
    #                 'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
    #                 'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
    #                 'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
    #                 'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
    #                 'LEFT JOIN "File" ON "File"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
    #                 'Left JOIN "CV_ElectronicFormat" ON "CV_ElectronicFormat"."Name"="File"."ElectronicFileFormatCV" '\
    #                 'WHERE "AttributeDataTypeCV"="File" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
    #                 .format(selectedType, selectedInstance, selectedAttribute)
    #
    #         result = self.session.execute(sql)
    #         # nameResult = list()
    #         complete_result = list()
    #         for row in result:
    #             # isExisting = False
    #             # for name in nameResult:
    #             #     if name == row.InstanceName:
    #             #         isExisting = True
    #             #         break
    #             # if not isExisting:
    #             #     nameResult.append(row.InstanceName)
    #             complete_result.append([row.InstanceName, row.ScenarioName,
    #                                     row.AttributeName, row.SourceName, row.MethodName,
    #                                     row.FileName, row.ElectronicFileFormatCV, excelPath, row.Description])
    #
    #         if excelPath != '':
    #             if complete_result.__len__() > 0:
    #                 self.isMatching_query(complete_result, "ElectronicFiles")
    #                 self.write2excel(complete_result, '4_ElectronicFiles', 14, 10, excelPath)
    #         return complete_result
    #     except Exception as  e:
    #         print e
    #         return []
    #         # raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportSeasonalSheet(self, selectedType='', selectedAttribute='', selectedInstance='', excelPath=''):
        '''
        This method is used to get data making SeasonalNumericValues_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedType == '' and selectedAttribute == '' and selectedInstance == '':
                sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName,SeasonName, SeasonNumericValue, SeasonNameCV ' \
                    'FROM "Attributes" '\
                    'Left JOIN "ObjectTypes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "SeasonalNumericValues" ON "SeasonalNumericValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE "AttributeDataTypeCV"="SeasonalNumericValues"'
            else:
                sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                  'ScenarioName,MethodName,SeasonName, SeasonNumericValue, SeasonNameCV ' \
                'FROM "Attributes" '\
                'Left JOIN "ObjectTypes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                'LEFT JOIN "SeasonalNumericValues" ON "SeasonalNumericValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                'WHERE "AttributeDataTypeCV"="SeasonalNumericValues" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.SeasonName, row.SeasonNameCV, row.SeasonNumericValue])
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "SeasonaNumericValues")
                    self.write2excel(complete_result, '4_SeasonaNumericValues', 11, 10, excelPath)
            return complete_result
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportTimeSeriesSheet(self, selectedType='', selectedAttribute='', selectedInstance='', excelPath=''):
        '''
        This method is used to get data making TimeSeries_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedType == '' and selectedAttribute == '' and selectedInstance == '':
                sql = 'SELECT ResourceType ObjectType, AttributeName, SourceName, InstanceName,YearType,' \
                      'ScenarioName,MethodName,AggregationStatisticCV, AggregationInterval, IntervalTimeUnitCV,' \
                      'IsRegular, NoDataValue, "TimeSeries"."Description" ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "TimeSeries" ON "TimeSeries"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE AttributeName!="ObjectInstances"  AND AttributeDataTypeCV="TimeSeries" '
            else:
                sql = 'SELECT ResourceType ObjectType, AttributeName, SourceName, InstanceName,YearType,' \
                      'ScenarioName,MethodName,AggregationStatisticCV, AggregationInterval, IntervalTimeUnitCV,' \
                      'IsRegular, NoDataValue, "TimeSeries"."Description" ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "TimeSeries" ON "TimeSeries"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE AttributeName!="ObjectInstances"  AND AttributeDataTypeCV="TimeSeries" ' \
                      ' AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                    .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.YearType, row.AggregationStatisticCV, row.AggregationInterval,
                                        row.IntervalTimeUnitCV, row.IsRegular, row.NoDataValue, row.Description])
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "TimeSeries")
                    self.write2excel(complete_result, '4_TimeSeries', 15, 10, excelPath)
            return complete_result
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportFreeTextSheet(self, selectedType='', selectedAttribute='', selectedInstance='', excelPath=''):
        '''
        This method is used to get data making FreeTextSheet.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedType == '' and selectedAttribute == ''and selectedInstance == '':
                sql = 'SELECT ResourceType, ObjectType, AttributeName, SourceName, InstanceName,FreeTextValue,' \
                      'ScenarioName,MethodName ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "FreeText" ON "FreeText"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE AttributeDataTypeCV="FreeText" '
            else:
                sql = 'SELECT ResourceType, ObjectType, AttributeName, SourceName, InstanceName,FreeTextValue,' \
                      'ScenarioName,MethodName ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "FreeText" ON "FreeText"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE AttributeDataTypeCV="FreeText" ' \
                      ' AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                    .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,row.FreeTextValue])
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "FreeText")
                    self.write2excel(complete_result, '4_FreeText', 12, 10, excelPath)
            return complete_result
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportMultiSheet(self, selectedType='', selectedAttribute='', selectedInstance='', excelPath=''):
        '''
        This method is used to get data making MultiAttributeSeries_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedType == '' and selectedInstance == '' and selectedAttribute == '':
                sql ="""
                    SELECT "ObjectTypes"."ObjectType",
                    "Instances"."InstanceName",ScenarioName,"Attributes"."AttributeName" AS MultiAttributeName,"Attributes".AttributeDataTypeCV,
                    SourceName,MethodName,
                    "AttributesColumns"."AttributeName" AS "AttributeName",
                    "AttributesColumns"."AttributeNameCV",
                    "AttributesColumns"."UnitNameCV" AS "AttributeNameUnitName",
                    "ValueOrder","DataValue"

                    FROM ResourceTypes

                    Left JOIN "ObjectTypes"
                    ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID"

                    -- Join the Object types to get their attributes
                    LEFT JOIN  "Attributes"
                    ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID"

                    -- Join the Attributes to get their Mappings
                    LEFT JOIN "Mappings"
                    ON Mappings.AttributeID= Attributes.AttributeID

                    -- Join the Mappings to get their Instances
                    LEFT JOIN "Instances"
                    ON "Instances"."InstanceID"="Mappings"."InstanceID"

                    -- Join the Mappings to get their ScenarioMappings
                    LEFT JOIN "ScenarioMappings"
                    ON "ScenarioMappings"."MappingID"="Mappings"."MappingID"

                    -- Join the ScenarioMappings to get their Scenarios
                    LEFT JOIN "Scenarios"
                    ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID"


                    -- Join the Scenarios to get their MasterNetworks
                    LEFT JOIN "MasterNetworks"
                    ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID"

                    -- Join the Mappings to get their Methods
                    LEFT JOIN "Methods"
                    ON "Methods"."MethodID"="Mappings"."MethodID"

                    -- Join the Mappings to get their Sources
                    LEFT JOIN "Sources"
                    ON "Sources"."SourceID"="Mappings"."SourceID"

                    -- Join the Mappings to get their DataValuesMappers
                    LEFT JOIN "ValuesMapper"
                    ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID"

                    -- Join the DataValuesMapper to get their MultiAttributeSeries
                    LEFT JOIN "MultiAttributeSeries"
                    ON "MultiAttributeSeries" ."ValuesMapperID"="ValuesMapper"."ValuesMapperID"


                    /*This is an extra join to get to each column name within the MultiColumn Array */

                    -- Join the MultiAttributeSeries to get to their specific DataValuesMapper, now called DataValuesMapperColumn
                    LEFT JOIN "ValuesMapper" As "ValuesMapperColumn"
                    ON "ValuesMapperColumn"."ValuesMapperID"="MultiAttributeSeries"."MappingID_Attribute"

                    -- Join the DataValuesMapperColumn to get back to their specific Mapping, now called MappingColumns
                    LEFT JOIN "Mappings" As "MappingColumns"
                    ON "MappingColumns"."ValuesMapperID"="ValuesMapperColumn"."ValuesMapperID"

                    -- Join the MappingColumns to get back to their specific Attribute, now called AttributeColumns
                    LEFT JOIN  "Attributes" AS "AttributesColumns"
                    ON "AttributesColumns"."AttributeID"="MappingColumns"."AttributeID"
                    /* Finishes here */

                    -- Join the MultiAttributeSeries to get access to their MultiAttributeSeriesValues
                    LEFT JOIN "MultiAttributeSeriesValues"
                    ON "MultiAttributeSeriesValues"."MultiAttributeSeriesID"="MultiAttributeSeries"."MultiAttributeSeriesID"

                    -- Select one InstanceName and restrict the query AttributeDataTypeCV that is MultiAttributeSeries
                    WHERE
                    "Attributes".AttributeDataTypeCV='MultiAttributeSeries'


                    Order By ScenarioName, AttributeName,ValueOrder asc

                    """
            else:
                sql ="""
                    SELECT "ObjectTypes"."ObjectType",
                    "Instances"."InstanceName",ScenarioName,"Attributes"."AttributeName" AS MultiAttributeName,"Attributes".AttributeDataTypeCV,
                    SourceName,MethodName,
                    "AttributesColumns"."AttributeName" AS "AttributeName",
                    "AttributesColumns"."AttributeNameCV",
                    "AttributesColumns"."UnitNameCV" AS "AttributeNameUnitName",
                    "ValueOrder","DataValue"

                    FROM ResourceTypes

                    Left JOIN "ObjectTypes"
                    ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID"

                    -- Join the Object types to get their attributes
                    LEFT JOIN  "Attributes"
                    ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID"

                    -- Join the Attributes to get their Mappings
                    LEFT JOIN "Mappings"
                    ON Mappings.AttributeID= Attributes.AttributeID

                    -- Join the Mappings to get their Instances
                    LEFT JOIN "Instances"
                    ON "Instances"."InstanceID"="Mappings"."InstanceID"

                    -- Join the Mappings to get their ScenarioMappings
                    LEFT JOIN "ScenarioMappings"
                    ON "ScenarioMappings"."MappingID"="Mappings"."MappingID"

                    -- Join the ScenarioMappings to get their Scenarios
                    LEFT JOIN "Scenarios"
                    ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID"


                    -- Join the Scenarios to get their MasterNetworks
                    LEFT JOIN "MasterNetworks"
                    ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID"

                    -- Join the Mappings to get their Methods
                    LEFT JOIN "Methods"
                    ON "Methods"."MethodID"="Mappings"."MethodID"

                    -- Join the Mappings to get their Sources
                    LEFT JOIN "Sources"
                    ON "Sources"."SourceID"="Mappings"."SourceID"

                    -- Join the Mappings to get their DataValuesMappers
                    LEFT JOIN "ValuesMapper"
                    ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID"

                    -- Join the DataValuesMapper to get their MultiAttributeSeries
                    LEFT JOIN "MultiAttributeSeries"
                    ON "MultiAttributeSeries" ."ValuesMapperID"="ValuesMapper"."ValuesMapperID"


                    /*This is an extra join to get to each column name within the MultiColumn Array */

                    -- Join the MultiAttributeSeries to get to their specific DataValuesMapper, now called DataValuesMapperColumn
                    LEFT JOIN "ValuesMapper" As "ValuesMapperColumn"
                    ON "ValuesMapperColumn"."ValuesMapperID"="MultiAttributeSeries"."MappingID_Attribute"

                    -- Join the DataValuesMapperColumn to get back to their specific Mapping, now called MappingColumns
                    LEFT JOIN "Mappings" As "MappingColumns"
                    ON "MappingColumns"."ValuesMapperID"="ValuesMapperColumn"."ValuesMapperID"

                    -- Join the MappingColumns to get back to their specific Attribute, now called AttributeColumns
                    LEFT JOIN  "Attributes" AS "AttributesColumns"
                    ON "AttributesColumns"."AttributeID"="MappingColumns"."AttributeID"
                    /* Finishes here */

                    -- Join the MultiAttributeSeries to get access to their MultiAttributeSeriesValues
                    LEFT JOIN "MultiAttributeSeriesValues"
                    ON "MultiAttributeSeriesValues"."MultiAttributeSeriesID"="MultiAttributeSeries"."MultiAttributeSeriesID"

                    -- Select one InstanceName and restrict the query AttributeDataTypeCV that is MultiAttributeSeries
                    WHERE
                    "Attributes".AttributeDataTypeCV='MultiAttributeSeries'
                    AND "ObjectTypeCV"="{}"
                    AND "InstanceNameCV"= "{}"
                    AND "Attributes"."AttributeNameCV" ="{}"


                    Order By ScenarioName, AttributeName,ValueOrder asc

                    """.format(selectedType, selectedInstance, selectedAttribute)


            result = self.session.execute(sql)

            '''Down Table(MultiVariableSeries_table Table) write'''
            complete_result = list()
            strAtrributName = ''
            valueOrder = None
            columnName = ''
            tempColumn = {}
            sourceName = ''
            i = 0
            currentrow = 0
            setNumber = 0
            for row in result:
                if row.MultiAttributeName == None or row.MultiAttributeName == "":
                    continue
                if strAtrributName != row.MultiAttributeName:
                    strAtrributName = row.MultiAttributeName
                    tempColumn[row.MultiAttributeName] = []
                    tempColumn[row.MultiAttributeName].append(row.MultiAttributeName)
                    columnName = row.MultiAttributeName
                if sourceName != row.ScenarioName:
                    sourceName = row.ScenarioName
                    setNumber = i
                    currentrow = 0
                if columnName != row.MultiAttributeName:
                    columnName = row.MultiAttributeName
                    currentrow = 0

                if row.MultiAttributeName in tempColumn[row.MultiAttributeName]:
                    index = tempColumn[row.MultiAttributeName].index(row.MultiAttributeName)
                    if index == 0:
                        complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName, row.MultiAttributeName, row.SourceName,
                                        row.MethodName, row.DataValue])
                        i += 1
                    else:
                        complete_result[setNumber + currentrow].append(row.DataValue)
                        currentrow += 1
                else:
                    currentrow = 0
                    tempColumn[row.MultiAttributeName].append(row.MultiAttributeName)
                    index = tempColumn[row.MultiAttributeName].index(row.MultiAttributeName)
                    if index == 0:
                        complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName, row.MultiAttributeName, row.SourceName,
                                        row.MethodName, row.DataValue])
                        i += 1
                    else:
                        complete_result[setNumber + currentrow].append(row.DataValue)
                        currentrow += 1
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "MultiVariableSeries")
                    self.write2excel(complete_result, '4_MultiVariableSeries', 17, 15, excelPath)
            # return complete_result

            bottom_table_result = complete_result

            '''Up Table(MultiAttributeName_column Table) write'''
            up_table_column_result = list()
            for key, columnItems in tempColumn.items():
                temList = list()
                temList.append(key)
                for item in columnItems:
                    temList.append(item)
                up_table_column_result.append(temList)
            if complete_result.__len__() > 0:
                self.write2excel(up_table_column_result, '4_MultiAttributeSeries', 17, 5, excelPath, 5)
            return up_table_column_result, bottom_table_result

        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)

    # def exportDualValuesSheet1(self, selectedResourceType, selectedNetwork, selectedScenarior, excelPath):
    #     '''
    #     This method is used to get data making DualValues_table.
    #     :param selectedResourceType: selected Model name
    #     :param selectedNetwork: selected Master Network name
    #     :param selectedScenarior: selected scenario Name
    #     :param excelPath: full path of excel file to export data
    #     :return: None
    #     '''
    #     try:
    #
    #         sql = 'SELECT AttributeName, SourceName, InstanceName,MasterNetworkName,ScenarioName,MethodName,dualvaluemeaningCV ' \
    #             'FROM "ResourceTypes" '\
    #             'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
    #             'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
    #             'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
    #             'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
    #             'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
    #             'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
    #             'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
    #             'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
    #             'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
    #             'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
    #             'LEFT JOIN "DualValues" ON "DualValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
    #             'LEFT JOIN "CV_DualValueMeaning" ON "CV_DualValueMeaning"."Name"= "DualValues"."dualvaluemeaningCV" '\
    #             'WHERE "Attributes"."AttributeDataTypeCV"="DualValues" AND "ResourceTypes"."ResourceTypeAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
    #             .format(selectedResourceType, selectedNetwork, selectedScenarior)
    #
    #         result = self.session.execute(sql)
    #         # nameResult = list()
    #         complete_result = list()
    #         for row in result:
    #             # isExisting = False
    #             # for name in nameResult:
    #             #     if name == row.InstanceName:
    #             #         isExisting = True
    #             #         break
    #             # if not isExisting:
    #             #     nameResult.append(row.InstanceName)
    #             complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
    #                                     row.AttributeName, row.SourceName, row.MethodName,
    #                                     row.dualvaluemeaningCV])
    #         if complete_result.__len__() > 0:
    #             self.isMatching_query(complete_result, "DualValues")
    #             self.write2excel(complete_result, '4_DualValues', 9, 10, excelPath)
    #     except Exception as  e:
    #         print e
    #         raise Exception('Error occurred in reading Data Structure.\n' + e.message)

    def exportTextConrolledSheet1(self, selectedResourceType='', selectedNetwork='', selectedScenarior='', excelPath=''):
        '''
        This method is used to get data making CategoricalValues_table.
        :param selectedResourceType: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedResourceType == '' and selectedNetwork == '' and selectedScenarior == '':
                sql = 'SELECT Attributes.AttributeName, ObjectType, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName,CategoricalValueCV ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "CategoricalValues" ON "CategoricalValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'LEFT JOIN "CV_Categorical" ON "CV_Categorical"."Name"= "CategoricalValues"."CategoricalValueCV" '\
                    'WHERE "Attributes"."AttributeDataTypeCV"="CategoricalValues"'
            else:
                sql = 'SELECT Attributes.AttributeName, ObjectType, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName,CategoricalValueCV ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "CategoricalValues" ON "CategoricalValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'LEFT JOIN "CV_Categorical" ON "CV_Categorical"."Name"= "CategoricalValues"."CategoricalValueCV" '\
                    'WHERE "Attributes"."AttributeDataTypeCV"="CategoricalValues" AND "ResourceTypes"."ResourceTypeAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                    .format(selectedResourceType, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.CategoricalValueCV])
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "CategoricalValues")
                    self.write2excel(complete_result, '4_CategoricalValues', 13, 10, excelPath)
            return complete_result
        except Exception as  e:
            raise Exception('Error occured in reading Data Structure.\n' + e.message)

    def exportNumericValuesheet1(self, selectedResourceType='', selectedNetwork='', selectedScenarior='', excelPath=''):
        '''
        This method is used to get data making Parameter.
        :param selectedResourceType: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedResourceType == '' and selectedNetwork == '' and selectedScenarior == '':
                sql = 'SELECT "ResourceTypes"."ResourceType", ObjectType,AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName, NumericValue ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID"'\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "NumericValues" ON "NumericValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE "Attributes"."AttributeDataTypeCV"="Parameter"'
            else:
                sql = 'SELECT "ResourceTypes"."ResourceType", ObjectType,AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName, NumericValue ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID"'\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "NumericValues" ON "NumericValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE "Attributes"."AttributeDataTypeCV"="Parameter" AND "ResourceTypes"."ResourceTypeAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                    .format(selectedResourceType, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.NumericValue])
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "Parameter")
                    self.write2excel(complete_result, '4_Parameter', 10, 10, excelPath)
            return complete_result
        except Exception as  e:
            print e
            raise Exception('Error occurer in reading Data Structure.\n' + e.message)
    # def exportElectronicFilesSheet1(self, selectedResourceType='', selectedNetwork='', selectedScenarior='', excelPath=''):
    #     '''
    #     This method is used to get data making ElectronicFiles.
    #     :param selectedResourceType: selected Model name
    #     :param selectedNetwork: selected Master Network name
    #     :param selectedScenarior: selected scenario Name
    #     :param excelPath: full path of excel file to export data
    #     :return: None
    #     '''
    #     try:
    #         if selectedResourceType =='' and selectedNetwork == '' and selectedScenarior == '':
    #             sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
    #                   'ScenarioName,MethodName,FileName, ElectronicFileFormatCV, "File"."Description" ' \
    #                 'FROM "ResourceTypes" '\
    #                 'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
    #                 'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
    #                 'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
    #                 'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
    #                 'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
    #                 'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
    #                 'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
    #                 'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
    #                 'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
    #                 'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
    #                 'LEFT JOIN "Files" ON "File"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
    #                 'Left JOIN "CV_ElectronicFormat" ON "CV_ElectronicFormat"."Name"="File"."ElectronicFileFormatCV" '\
    #                 'WHERE "Attributes"."AttributeDataTypeCV"="File"'
    #         else:
    #             sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
    #                   'ScenarioName,MethodName,FileName, ElectronicFileFormatCV, "File"."Description" ' \
    #                 'FROM "ResourceTypes" '\
    #                 'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
    #                 'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
    #                 'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
    #                 'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
    #                 'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
    #                 'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
    #                 'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
    #                 'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
    #                 'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
    #                 'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
    #                 'LEFT JOIN "Files" ON "File"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
    #                 'Left JOIN "CV_ElectronicFormat" ON "CV_ElectronicFormat"."Name"="File"."ElectronicFileFormatCV" '\
    #                 'WHERE "Attributes"."AttributeDataTypeCV"="File" AND "ResourceTypes"."ResourceTypeAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
    #                 .format(selectedResourceType, selectedNetwork, selectedScenarior)
    #
    #         result = self.session.execute(sql)
    #         # nameResult = list()
    #         complete_result = list()
    #         for row in result:
    #             # isExisting = False
    #             # for name in nameResult:
    #             #     if name == row.InstanceName:
    #             #         isExisting = True
    #             #         break
    #             # if not isExisting:
    #             #     nameResult.append(row.InstanceName)
    #             complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
    #                                     row.AttributeName, row.SourceName, row.MethodName,
    #                                     row.FileName, row.ElectronicFileFormatCV, excelPath, row.Description])
    #         if excelPath != '':
    #             if complete_result.__len__() > 0:
    #                 self.isMatching_query(complete_result, "ElectronicFiles")
    #                 self.write2excel(complete_result, '4_ElectronicFiles', 14, 10, excelPath)
    #         return complete_result
    #     except Exception as  e:
    #         pass
    #         # print e
    #         # raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportSeasonalSheet1(self, selectedResourceType='', selectedNetwork='', selectedScenarior='', excelPath=''):
        '''
        This method is used to get data making SeasonalParameter.
        :param selectedResourceType: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedResourceType == '' and selectedNetwork == '' and selectedScenarior == '':
                sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName,SeasonName, SeasonNumericValue, SeasonNameCV ' \
                    'FROM "Attributes" '\
                    'Left JOIN "ObjectTypes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "SeasonalNumericValues" ON "SeasonalNumericValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE "AttributeDataTypeCV"="SeasonaNumericValues"'
            else:
                sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                      'ScenarioName,MethodName,SeasonName, SeasonNumericValue, SeasonNameCV ' \
                    'FROM "Attributes" '\
                    'Left JOIN "ObjectTypes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "SeasonalNumericValues" ON "SeasonalNumericValues"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE "AttributeDataTypeCV"="SeasonaNumericValues" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                    .format(selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.SeasonName, row.SeasonNameCV, row.SeasonNumericValue])
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "SeasonaNumericValues")
                    self.write2excel(complete_result, '4_SeasonaNumericValues', 11, 10, excelPath)
            return complete_result
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportTimeSeriesSheet1(self, selectedResourceType='', selectedNetwork='', selectedScenarior='', excelPath=''):
        '''
        This method is used to get data making TimeSeries.
        :param selectedResourceType: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedResourceType == '' and selectedNetwork == '' and selectedScenarior == '':
                sql = 'SELECT ResourceType ObjectType, AttributeName, SourceName, InstanceName,YearType,' \
                      'ScenarioName,MethodName,AggregationStatisticCV, AggregationInterval, IntervalTimeUnitCV,' \
                      'IsRegular, NoDataValue, "TimeSeries"."Description" ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "TimeSeries" ON "TimeSeries"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE AttributeName!="ObjectInstances"  AND AttributeDataTypeCV="TimeSeries" '
            else:
                sql = 'SELECT ResourceType ObjectType, AttributeName, SourceName, InstanceName,YearType,' \
                      'ScenarioName,MethodName,AggregationStatisticCV, AggregationInterval, IntervalTimeUnitCV,' \
                      'IsRegular, NoDataValue, "TimeSeries"."Description" ' \
                    'FROM "ResourceTypes" '\
                    'Left JOIN "ObjectTypes" ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID" '\
                    'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                    'Left JOIN "Mappings" ON "Mappings"."AttributeID"= "Attributes"."AttributeID" '\
                    'Left JOIN "ValuesMapper" ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID" '\
                    'Left JOIN "ScenarioMappings" ON "ScenarioMappings"."MappingID"="Mappings"."MappingID" '\
                    'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID" '\
                    'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                    'Left JOIN "Methods" ON "Methods"."MethodID"="Mappings"."MethodID" '\
                    'Left JOIN "Sources" ON "Sources"."SourceID"="Mappings"."SourceID" '\
                    'Left JOIN "Instances" ON "Instances"."InstanceID"="Mappings"."InstanceID" '\
                    'LEFT JOIN "TimeSeries" ON "TimeSeries"."ValuesMapperID" = "ValuesMapper"."ValuesMapperID" '\
                    'WHERE AttributeName!="ObjectInstances"  AND AttributeDataTypeCV="TimeSeries" ' \
                       'AND "ResourceTypeAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                    .format(selectedResourceType, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.YearType, row.AggregationStatisticCV, row.AggregationInterval,
                                        row.IntervalTimeUnitCV, row.IsRegular, row.NoDataValue, row.Description])
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "TimeSeries")
                    self.write2excel(complete_result, '4_TimeSeries', 15, 10, excelPath)
            return complete_result
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)

    def exportMultiSheet1(self, selectedResourceType='', selectedNetwork='', selectedScenarior='', excelPath=''):
        '''
        This method is used to get data making MultiVariableSeries.
        :param selectedResourceType: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            if selectedResourceType == '' and selectedNetwork == '' and selectedScenarior == '':
                sql = """
                        SELECT "ObjectTypes"."ObjectType",
                        "Instances"."InstanceName",ScenarioName,"Attributes"."AttributeName" AS MultiAttributeName,"Attributes".AttributeDataTypeCV,
                        SourceName,MethodName,
                        "AttributesColumns"."AttributeName" AS "AttributeName",
                        "AttributesColumns"."AttributeNameCV",
                        "AttributesColumns"."UnitNameCV" AS "AttributeNameUnitName",
                        "ValueOrder","DataValue"

                        FROM ResourceTypes

                        Left JOIN "ObjectTypes"
                        ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID"

                        -- Join the Object types to get their attributes
                        LEFT JOIN  "Attributes"
                        ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID"

                        -- Join the Attributes to get their Mappings
                        LEFT JOIN "Mappings"
                        ON Mappings.AttributeID= Attributes.AttributeID

                        -- Join the Mappings to get their Instances
                        LEFT JOIN "Instances"
                        ON "Instances"."InstanceID"="Mappings"."InstanceID"

                        -- Join the Mappings to get their ScenarioMappings
                        LEFT JOIN "ScenarioMappings"
                        ON "ScenarioMappings"."MappingID"="Mappings"."MappingID"

                        -- Join the ScenarioMappings to get their Scenarios
                        LEFT JOIN "Scenarios"
                        ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID"


                        -- Join the Scenarios to get their MasterNetworks
                        LEFT JOIN "MasterNetworks"
                        ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID"

                        -- Join the Mappings to get their Methods
                        LEFT JOIN "Methods"
                        ON "Methods"."MethodID"="Mappings"."MethodID"

                        -- Join the Mappings to get their Sources
                        LEFT JOIN "Sources"
                        ON "Sources"."SourceID"="Mappings"."SourceID"

                        -- Join the Mappings to get their DataValuesMappers
                        LEFT JOIN "ValuesMapper"
                        ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID"

                        -- Join the DataValuesMapper to get their MultiAttributeSeries
                        LEFT JOIN "MultiAttributeSeries"
                        ON "MultiAttributeSeries" ."ValuesMapperID"="ValuesMapper"."ValuesMapperID"


                        /*This is an extra join to get to each column name within the MultiColumn Array */

                        -- Join the MultiAttributeSeries to get to their specific DataValuesMapper, now called DataValuesMapperColumn
                        LEFT JOIN "ValuesMapper" As "ValuesMapperColumn"
                        ON "ValuesMapperColumn"."ValuesMapperID"="MultiAttributeSeries"."MappingID_Attribute"

                        -- Join the DataValuesMapperColumn to get back to their specific Mapping, now called MappingColumns
                        LEFT JOIN "Mappings" As "MappingColumns"
                        ON "MappingColumns"."ValuesMapperID"="ValuesMapperColumn"."ValuesMapperID"

                        -- Join the MappingColumns to get back to their specific Attribute, now called AttributeColumns
                        LEFT JOIN  "Attributes" AS "AttributesColumns"
                        ON "AttributesColumns"."AttributeID"="MappingColumns"."AttributeID"
                        /* Finishes here */

                        -- Join the MultiAttributeSeries to get access to their MultiAttributeSeriesValues
                        LEFT JOIN "MultiAttributeSeriesValues"
                        ON "MultiAttributeSeriesValues"."MultiAttributeSeriesID"="MultiAttributeSeries"."MultiAttributeSeriesID"

                        -- Select one InstanceName and restrict the query AttributeDataTypeCV that is MultiAttributeSeries
                        WHERE
                        "Attributes".AttributeDataTypeCV='MultiAttributeSeries'

                         Order By ScenarioName, AttributeName,ValueOrder asc

                        """
            else:
                sql = """
                        SELECT "ObjectTypes"."ObjectType",
                        "Instances"."InstanceName",ScenarioName,"Attributes"."AttributeName" AS MultiAttributeName,"Attributes".AttributeDataTypeCV,
                        SourceName,MethodName,
                        "AttributesColumns"."AttributeName" AS "AttributeName",
                        "AttributesColumns"."AttributeNameCV",
                        "AttributesColumns"."UnitNameCV" AS "AttributeNameUnitName",
                        "ValueOrder","DataValue"

                        FROM ResourceTypes

                        Left JOIN "ObjectTypes"
                        ON "ObjectTypes"."ResourceTypeID"="ResourceTypes"."ResourceTypeID"

                        -- Join the Object types to get their attributes
                        LEFT JOIN  "Attributes"
                        ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID"

                        -- Join the Attributes to get their Mappings
                        LEFT JOIN "Mappings"
                        ON Mappings.AttributeID= Attributes.AttributeID

                        -- Join the Mappings to get their Instances
                        LEFT JOIN "Instances"
                        ON "Instances"."InstanceID"="Mappings"."InstanceID"

                        -- Join the Mappings to get their ScenarioMappings
                        LEFT JOIN "ScenarioMappings"
                        ON "ScenarioMappings"."MappingID"="Mappings"."MappingID"

                        -- Join the ScenarioMappings to get their Scenarios
                        LEFT JOIN "Scenarios"
                        ON "Scenarios"."ScenarioID"="ScenarioMappings"."ScenarioID"


                        -- Join the Scenarios to get their MasterNetworks
                        LEFT JOIN "MasterNetworks"
                        ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID"

                        -- Join the Mappings to get their Methods
                        LEFT JOIN "Methods"
                        ON "Methods"."MethodID"="Mappings"."MethodID"

                        -- Join the Mappings to get their Sources
                        LEFT JOIN "Sources"
                        ON "Sources"."SourceID"="Mappings"."SourceID"

                        -- Join the Mappings to get their DataValuesMappers
                        LEFT JOIN "ValuesMapper"
                        ON "ValuesMapper"."ValuesMapperID"="Mappings"."ValuesMapperID"

                        -- Join the DataValuesMapper to get their MultiAttributeSeries
                        LEFT JOIN "MultiAttributeSeries"
                        ON "MultiAttributeSeries" ."ValuesMapperID"="ValuesMapper"."ValuesMapperID"


                        /*This is an extra join to get to each column name within the MultiColumn Array */

                        -- Join the MultiAttributeSeries to get to their specific DataValuesMapper, now called DataValuesMapperColumn
                        LEFT JOIN "ValuesMapper" As "ValuesMapperColumn"
                        ON "ValuesMapperColumn"."ValuesMapperID"="MultiAttributeSeries"."MappingID_Attribute"

                        -- Join the DataValuesMapperColumn to get back to their specific Mapping, now called MappingColumns
                        LEFT JOIN "Mappings" As "MappingColumns"
                        ON "MappingColumns"."ValuesMapperID"="ValuesMapperColumn"."ValuesMapperID"

                        -- Join the MappingColumns to get back to their specific Attribute, now called AttributeColumns
                        LEFT JOIN  "Attributes" AS "AttributesColumns"
                        ON "AttributesColumns"."AttributeID"="MappingColumns"."AttributeID"
                        /* Finishes here */

                        -- Join the MultiAttributeSeries to get access to their MultiAttributeSeriesValues
                        LEFT JOIN "MultiAttributeSeriesValues"
                        ON "MultiAttributeSeriesValues"."MultiAttributeSeriesID"="MultiAttributeSeries"."MultiAttributeSeriesID"

                        -- Select one InstanceName and restrict the query AttributeDataTypeCV that is MultiAttributeSeries
                        WHERE
                        "Attributes".AttributeDataTypeCV='MultiAttributeSeries'
                        AND "ResourceTypeAcronym"="{}"
                        AND "MasterNetworkName"= "{}"
                        AND "ScenarioName" ="{}"


                         Order By ScenarioName, AttributeName,ValueOrder asc

                        """.format(selectedResourceType, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)

            '''Down Table(MultiVariableSeries_table Table) write'''
            complete_result = list()
            strAtrributName = ''
            valueOrder = None
            AttributeName = ''
            tempColumn = {}
            sourceName = ''
            i = 0
            currentrow = 0
            setNumber = 0
            for row in result:
                if row.AttributeName == None or row.AttributeName == "":
                    continue
                if strAtrributName != row.AttributeName:
                    strAtrributName = row.AttributeName
                    tempColumn[row.AttributeName] = []
                    tempColumn[row.AttributeName].append(row.AttributeName)
                    AttributeName = row.AttributeName
                if sourceName != row.ScenarioName:
                    sourceName = row.ScenarioName
                    setNumber = i
                    currentrow = 0
                if AttributeName != row.AttributeName:
                    AttributeName = row.AttributeName
                    currentrow = 0

                if row.AttributeName in tempColumn[row.AttributeName]:
                    index = tempColumn[row.AttributeName].index(row.AttributeName)
                    if index == 0:
                        complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName, row.AttributeName, row.SourceName,
                                        row.MethodName, row.DataValue])
                        i += 1
                    else:
                        complete_result[setNumber + currentrow].append(row.DataValue)
                        currentrow += 1
                else:
                    currentrow = 0
                    tempColumn[row.AttributeName].append(row.AttributeName)
                    index = tempColumn[row.AttributeName].index(row.AttributeName)
                    if index == 0:
                        complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName, row.AttributeName, row.SourceName,
                                        row.MethodName, row.DataValue])
                        i += 1
                    else:
                        complete_result[setNumber + currentrow].append(row.DataValue)
                        currentrow += 1
            if excelPath != '':
                if complete_result.__len__() > 0:
                    self.isMatching_query(complete_result, "MultiAttributeSeries")
                    self.write2excel(complete_result, '4_MultiAttributeSeries', 17, 15, excelPath)
            return complete_result

            '''Up Table(AttributeName_column Table) write'''
            column_result = list()
            for key, columnItems in tempColumn.items():
                temList = list()
                temList.append(key)
                for item in columnItems:
                    temList.append(item)
                column_result.append(temList)
            if complete_result.__len__() > 0:
                self.write2excel(column_result, '4_MultiAttributeSeries', 17, 5, excelPath, 5)

        except Exception as  e:
            print e
            raise Exception('Error occured in reading Data Structure.\n' + e.message)


    def write2excel(self, complete_result, xlsx_sheet_name, xls_sheet_num, startRowNum, excelPath, columnNum = 0):
        '''
        This method is used to write data specific sheet.
        :param complete_result: data to write within sheet
        :param xlsx_sheet_name: sheet name
        :param xls_sheet_num: sheet number to write
        :param startRowNum: start roe number to write
        :param excelPath: full path of excel file to export data
        :param columnNum: column number
        :return: None
        '''
        try:
            if excelPath.split('.')[-1] == 'xls':
                wb = open_workbook(excelPath)
                # workbook = copy(wb)
                workbook = wb
                try:
                    sheet = workbook.get_sheet(xls_sheet_num)
                except:
                    raise Exception("Please select a valid Excel File")
                for row_id, row in enumerate(complete_result):
                    for col_id, cell in enumerate(row):
                        sheet.write(row_id + startRowNum - 1, col_id + columnNum, cell)
                workbook.save(excelPath)
            else:
                book2 = load_workbook(excelPath)
                try:
                    sheet = book2.get_sheet_by_name(xlsx_sheet_name)
                except:
                    raise Exception('Output Sheet {} not found in Excel File \n Please select valid Excel File'.format(xlsx_sheet_name))

                for row_id, row in enumerate(complete_result):
                    for col_id, cell in enumerate(row):
                        sheet.cell(row=row_id + startRowNum, column=col_id + columnNum + 1, value=cell)
                book2.save(excelPath)

        except Exception as e:
            print e
            raise Exception(e.message)

    def isMatching_query(self, resultList, stateQuery):
        if (len(resultList) < 1):
            raise Exception("Nothing found in the {} selected WAMDAM database to match your request".format(stateQuery))

