

"""
    The shapefileExporter.py is a python script use to reshape
    a shapefile data and write the output result in the appropriate
    excel sheets for WaMDaM template. it is made up of:

    - shapefileExporter():  This is the function which shapes the  excel file
                by looking for distinct attribute types and doing a cross
                multiplication with the file properties and the attributes
                values using the corresponding object code.
                Warning: if format of file changes, script will fail to work

    - write_data(): This is a helper function used in accomplishing shapefileExporter()
                    all it does is write the data to the appropriate sheets
"""

# The xlrd library is used for only reading data in excel files. its is quite fast at that
import xlrd as excel

# This library is used here to write data to an excel file
from openpyxl import load_workbook



class shapefileExporter():
    """
    This class gets data from an xml file given by the user
    and stores in the Wamdam database
    """

    def __init__(self):
        pass
        # self.shapefileExporter(workbook)


    def shapefileExporter(self, workbook, input_sheet='Shapefile_input'):
        """
            This function parses the input file for shapefileExporter and get all required
            matrices and control data. the user must make sure that the excel file
            has sheetnames ('DualValues', 'NumericValues' and 'DescriptorValues')
            :param workbook: holds the excel file
            :param input_sheet: holds the input sheet
            :return: None
        """
        book = excel.open_workbook(workbook)
        try:
            sheet = book.sheet_by_name(input_sheet)
        except:
            raise Exception('Input sheet {} not found in the provided Excel file \n Please select valid excel file'.format(input_sheet))

        rows = [rows for rows in sheet.get_rows()]


        sources = [cell for cell in rows[1][3:]]

        methods = [cell for cell in rows[3][3:]]

        attributes = [str(cell.value) for cell in rows[5][3:]]

        attrib_code = [cell for cell in rows[7][3:]]

        left_data = [row[0:3] for row in rows[8:]]


        right_data = [row[3:] for row in rows[8:]]

        book2 = load_workbook(workbook)

        max_row = [9, 9, 9, 9]
        try:
            for ida, attrib in enumerate(attributes):

                if attrib == '4_FreeText':
                    # write_data('4_DualValues', book2, ida, attrib_code, sources, methods, left_data, right_data, max_row[1])
                    max_row[1] = self.write_data('4_FreeText', book2, ida, attrib_code, sources, methods, left_data, right_data, max_row[1] + 1)

                elif attrib == 'NumericValues':
                    # write_data('4_NumericValues', book2, ida, attrib_code, sources, methods, left_data, right_data, max_row[2])
                    max_row[2] = self.write_data('4_NumericValues', book2, ida, attrib_code, sources, methods, left_data, right_data, max_row[2] + 1)

                elif attrib == '4_CategoricalValues':
                    # write_data('4_DescriptorValues', book2, ida, attrib_code, sources, methods, left_data, right_data, max_row[3])
                    max_row[3] = self.write_data('4_CategoricalValues', book2, ida, attrib_code, sources, methods, left_data, right_data, max_row[3] + 1)
        except Exception as e:
            raise Exception(e.message)


        book2.save(workbook)

    def write_data(self, sheet_name, workbook, ida, attrib_code, sources, methods, left_data, right_data, max_row):
        """
            This is a helper function for shapefileExporter to write data in
            the appropriate excel sheet.
            :param sheet_name: Sheet name sent from shapefileExporter() function
            :param workbook: The excel workbook in question
            :param ida: Attribute index (integer)
            :param attrib_code: Attribute code (eg 'DAM_ID_DamUtah')
            :param left_data: utahDams properties
            :param right_data: Attribute values
            :return:
        """
        try:
            sheet = workbook.get_sheet_by_name(sheet_name)
        except:
            raise Exception(
                'Output sheet {} not found in Excel File\nPlease select a valid excel file'.format(sheet_name))
        temp = list()
        for idd, data in enumerate(left_data):
            content = data[:]
            content.append(attrib_code[ida])
            content.append(sources[ida])
            content.append(methods[ida])
            content += [right_data[idd][ida]]
            temp.append(content)

        # writing parsing result to appropriate excel sheet

        for rowID, row in enumerate(temp):
            for colID, cell in enumerate(row):
                try:
                    sheet.cell(row=rowID + max_row, column=colID + 1, value=unicode(cell.value))
                except Exception as e:
                    if e.args[0] == 'ascii':
                        # raise Exception(e)
                        raise Exception("There is word '" + e.args[
                            1] + "' that is not a unicode type in the sheet. Please fix the word.")
                    raise Exception(e)
        return rowID + max_row


