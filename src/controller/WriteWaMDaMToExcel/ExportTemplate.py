# maybe write the script here? then send it to the viewer dlg_ExportScenarioDataToHydra  ?

from controller.wamdamAPI.GetResourceStructure import GetResourceStructure
from controller.wamdamAPI.GetDataValues import GetDataValues
from controller.ConnectDB_ParseExcel import DB_Setup
from controller.ConnectDB_ParseExcel import SqlAlchemy as sq

import xlsxwriter, os, Global

from openpyxl import load_workbook

import pandas as pd

class ExportTemplate (object):
	def __init__(self, full_path):

		workbook1 = xlsxwriter.Workbook(full_path)
		workbook1.add_worksheet('1.1_Organiz&People')
		workbook1.add_worksheet('1.2_Sources&Methods')
		workbook1.add_worksheet('2.1_ResourceTypes&ObjectTypes')
		workbook1.add_worksheet('2.2_Attributes')
		workbook1.add_worksheet('3.1_Networks&Scenarios')
		workbook1.add_worksheet('3.2_Nodes')
		workbook1.add_worksheet('3.3_Links')
		workbook1.add_worksheet('4_NumericValues')
		workbook1.add_worksheet('4_FreeText')
		workbook1.add_worksheet('4_CategoricalValues')
		workbook1.add_worksheet('4_SeasonalNumericValues')
		workbook1.add_worksheet('4_TimeSeries')
		workbook1.add_worksheet('4_TimeSeriesValues')
		workbook1.add_worksheet('4_MultiAttributeSeries')
		workbook1.add_worksheet('ObjectCategory')
		workbook1.add_worksheet('AttributeCategory')
		workbook1.add_worksheet('InstanceCategory')
		workbook1.close()
		##########################################################
		
		self.workbook = load_workbook(full_path)
		self.file_path = full_path

		self.pandas_excel_writer = pd.ExcelWriter(full_path, engine='xlsxwriter')

	def exportOrganizations(self, organization_list):
		# df = pd.DataFrame(organization_list)
		organization_list.to_excel(self.pandas_excel_writer, sheet_name="1.1_Organiz&People", startrow=2, startcol=0, header=True,
					index=False)
		# self.write2excel(organization_list, 2, '1.1_Organiz&People')

	def exportPeople(self, people_list):
		df = pd.DataFrame(people_list)
		df.to_excel(self.pandas_excel_writer, sheet_name="1.1_Organiz&People", startrow=9, startcol=0, header=True,
					index=False)
		# self.write2excel(people_list, 9, '1.1_Organiz&People')

	def exportSources(self, source_list):
		# df = pd.DataFrame(source_list)
		source_list.to_excel(self.pandas_excel_writer, sheet_name="1.2_Sources&Methods", startrow=8, startcol=0, header=True,
					index=False)
		# self.write2excel(source_list, 8, '1.2_Sources&Methods')

	def exportMethods(self, method_list):
		df = pd.DataFrame(method_list)
		df.to_excel(self.pandas_excel_writer, sheet_name="1.2_Sources&Methods", startrow=19, startcol=0, header=True,
					index=False)
		# self.write2excel(method_list, 19, '1.2_Sources&Methods')

	def exportMasterNetwork(self, network_data_result):

		network_data_result.to_excel(self.pandas_excel_writer, sheet_name="3.1_Networks&Scenarios", startrow=8, startcol=0, header=True,
					index=False)
		# scenario_data_result.to_excel(self.pandas_excel_writer, sheet_name="3.1_Networks&Scenarios", startrow=20,
		# 							 startcol=1, header=True,
		# 							 index=False)
		# self.write2excel(network_data_result, 10, '3.1_Networks&Scenarios')

	def exportScenario(self, scenarios_data_result):
		# df = pd.DataFrame(scenarios_data_result)
		scenarios_data_result.to_excel(self.pandas_excel_writer, sheet_name="3.1_Networks&Scenarios", startrow=18, startcol=0,
					header=True,
					index=False)
		# self.write2excel(scenarios_data_result, 20, '3.1_Networks&Scenarios')

	def exportNodes(self, nodes_data_result):
		# df = pd.DataFrame(nodes_data_result)
		nodes_data_result.to_excel(self.pandas_excel_writer, sheet_name="3.2_Nodes", startrow=8, startcol=0, header=True, index=False)
		# self.write2excel(nodes_data_result, 11, '3.2_Nodes')

	def exportLinkes(self, links_frame_result):
		# df = pd.DataFrame(links_data_result)
		links_frame_result.to_excel(self.pandas_excel_writer, sheet_name="3.3_Links", startrow=8, startcol=0, header=True,
					index=False)
		# self.write2excel(links_data_result, 11, '3.3_Links')

	def exportObjecttypes(self, data_result):
		df = pd.DataFrame(data_result)
		df.to_excel(self.pandas_excel_writer, sheet_name="2.1_ResourceTypes&ObjectTypes", header=True,
					index=False)
		# self.write2excel(data_result, 18, '2.1_ResourceTypes&ObjectTypes')

	def exportResourcesType(self, resources_result , objects_result):
		if not resources_result.empty:
			resources_result.to_excel(self.pandas_excel_writer, sheet_name="2.1_ResourceTypes&ObjectTypes", startrow=8, startcol=0,
					header=True,
					index=False)
		if not objects_result.empty:
			objects_result.to_excel(self.pandas_excel_writer, sheet_name="2.1_ResourceTypes&ObjectTypes", startrow=16, startcol=0,
					header=True,
					index=False)
		# self.write2excel(resources_result, 10, '2.1_ResourceTypes&ObjectTypes')

	def exportAttributes(self, attributes_result):
		# df = pd.DataFrame(attributes_result)
		attributes_result.to_excel(self.pandas_excel_writer, sheet_name="2.2_Attributes", startrow=9, startcol=0,
					header=True,
					index=False)
		# self.write2excel(attributes_result, 11, '2.2_Attributes')

	# def exportDualValues(self, selectedResourceType, selectedMasterNetworkName, selectedScenarioName):
	# 	self.getDataValues.exportDualValuesSheet1(selectedResourceType, selectedMasterNetworkName, selectedScenarioName, self.file_path)

	def exportFreeText(self, data_list):
		df = data_list
		df.to_excel(self.pandas_excel_writer, sheet_name="4_FreeText", startrow=8, startcol=0,
					header=True,
					index=False)

		# self.write2excel(data_list, 3, '4_FreeText')

	def exportNumericValue(self, data_list):
		df = data_list
		df.to_excel(self.pandas_excel_writer, sheet_name="4_NumericValues", startrow=8, startcol=0,
					header=True,
					index=False)
		# self.write2excel(data_list, 10, '4_NumericValues')

	def exportElectronicFiles(self, data_list):
		df = pd.DataFrame(data_list)
		df.to_excel(self.pandas_excel_writer, sheet_name="4_ElectronicFiles", startrow=8, startcol=0,
					header=True,
					index=False)
		# self.write2excel(data_list, 14, '4_ElectronicFiles')

	def exportSeasonal(self, data_list):

		# data_list
		# sort the data here

		# df = pd.DataFrame(data_list)
		data_list.to_excel(self.pandas_excel_writer, sheet_name="4_SeasonalNumericValues", startrow=8, startcol=0,
					header=True,
					index=False)
		# self.write2excel(data_list, 11, '4_SeasonalNumericValues')


	def exportTimeSeries(self, data_list):
		# df = pd.DataFrame(data_list)
		data_list.to_excel(self.pandas_excel_writer, sheet_name="4_TimeSeries", startrow=8, startcol=0,
					header=True,
					index=False)



	def exportTimeSeriesValues(self, data_list):
		df = data_list
		df.to_excel(self.pandas_excel_writer, sheet_name="4_TimeSeriesValues", startrow=8, startcol=0,
					header=True,
					index=False)


	def exportCategoricalValues(self, data_list):
		df = pd.DataFrame(data_list)
		df.to_excel(self.pandas_excel_writer, sheet_name="4_CategoricalValues", startrow=8, startcol=0,
					header=True,
					index=False)
		# self.write2excel(data_list, 13, '4_CategoricalValues')

	def exportMulti(self, up_table_column_result, bottom_table_result):
		# df = pd.DataFrame(bottom_table_result)
		up_table_column_result.to_excel(self.pandas_excel_writer, sheet_name="4_MultiAttributeSeries", startrow=3, startcol=5,
					header=True,
					index=False)

		# df = pd.DataFrame(up_table_column_result)
		bottom_table_result.to_excel(self.pandas_excel_writer, sheet_name="4_MultiAttributeSeries", startrow=17, startcol=0,
					header=True,
					index=False)
		# self.write2excel(up_table_column_result, 2, '4_MultiAttributeSeries')
		# self.write2excel(bottom_table_result, 20, '4_MultiAttributeSeries')

	def file_save(self):
		self.pandas_excel_writer.save()

	def write2excel(self, resutl_list, startRow, sheet_name):
		try:
			try:
				sheet = self.workbook.get_sheet_by_name(sheet_name)
			except:
				raise Exception("Please select a valid Excel File")
			for row_id, row in enumerate(resutl_list):
				for col_id, cell in enumerate(row):
					sheet.cell(row=row_id + startRow, column=col_id + 1, value=cell)

				print "writing row id: " + str(row_id)

				self.workbook.save(self.file_path)

		except Exception as e:
			print e
			raise Exception(e.message)