
import wx
import WaMDaMWizard
from controller.wamdamAPI.GetDataStructure import GetDataStructure
from controller.wamdamAPI.GetMetadata import GetMetadata
from controller.wamdamAPI.GetInstances import GetInstances
from controller.wamdamAPI.GetAllValuesByScenario import GetAllValuesByScenario
from controller.wamdamAPI.ExportTemplate import ExportTemplate
from xlrd import open_workbook
# from xlutils.copy import copy
from Messages_forms.msg_somethigWrong import msg_somethigWrong
# This library is used here to write data to an excel file
from openpyxl import load_workbook
import xlsxwriter
from controller.ConnectDB_ParseExcel import DB_Setup

# Implementing dlg_ExportScenarioDataToRwise
class dlg_ExportScenarioDataToHydra( WaMDaMWizard.dlg_ExportScenarioDataToHydra ):
	def __init__( self, parent ):
		WaMDaMWizard.dlg_ExportScenarioDataToHydra.__init__( self, parent )
		self.sheetNames = ['1.1_Organiz&People', '1.2_Sources&Methods',
						   '2.1_ResourceTypes&ObjectTypes', '2.2_Attributes', '3.1_Networks&Scenarios',
						   '3.2_Nodes', '3.3_Links', '4_CategoricalValues', '4_NumericVaules', '4_FreeText',
						   '4_SeasonalNumericValues', '4_ElectronicFiles', '4_TimeSeries', '4_MultiAttributeSeries']
		self.path = ''
		try:
			if not DB_Setup().get_session():
				msg = "\n\nWarning: Please connect to sqlite first."
				raise Exception(msg)
			''' init combo model'''
			self.dataStructure = GetDataStructure()
			self.datasets = self.dataStructure.getResourceTypes()
			list_acromy = list()
			for row in self.datasets:
				list_acromy.append(row.ResourceTypeAcronym)
			if list_acromy.__len__() > 0:
				self.comboBox_selectModel.SetItems(list_acromy)
		except Exception as e:
			message = msg_somethigWrong(None, msg=e.message)
			message.Show()
			self.Destroy()

	# Handlers for dlg_ExportScenarioDataToRwise events.
	def comboBox_selectModelOnCombobox( self, event ):
		selectedDataset = self.comboBox_selectModel.Value
		result = GetInstances()
		gettedMasterNetworkNames, data = result.getMasterNetwork(selectedDataset)
		if gettedMasterNetworkNames.__len__() <= 0:
			return
		self.comboBox_selectNetwork.SetItems(gettedMasterNetworkNames)
	
	def comboBox_selectNetworkOnCombobox( self, event ):
		selectedMasterNetworkName = self.comboBox_selectNetwork.Value
		result = GetInstances()
		gettedScenarioNames, data = result.getScenario(self.comboBox_selectModel.Value, selectedMasterNetworkName)
		if gettedScenarioNames.__len__() <= 0:
			return
		self.comboBox_selectScenario.SetItems(gettedScenarioNames)

	
	def comboBox_selectScenarioOnCombobox( self, event ):
		# TODO: Implement comboBox_selectScenarioOnCombobox
		pass
	
	def DirectoryPicker_ExportToRwiseOnFileChanged( self, event ):
		valid_extension = ['xls','xlsx']
		self.path = self.DirectoryPicker_ExportToRwise.GetPath()

		# Because use directory picker dialog, add file name.
		# Create file name to export containing "ResourceTypeAcronum" and "NetworkName".
		selectedDataset = self.comboBox_selectModel.Value
		selectedMasterNetworkName = self.comboBox_selectNetwork.Value
		file_name_to_export = "{}_{}".format(selectedDataset, selectedMasterNetworkName)
		self.path = "{}/{}.xlsx".format(self.path, file_name_to_export)

		# if not (self.path.split('.')[-1] in valid_extension):
		# 	message = msg_somethigWrong(None, msg="Please select a valid Excel File")
		# 	message.Show()
		# 	return
		print 'This is working just fine...'
		print self.path
	
	def btn_Export_ScenarioDataOnButtonClick( self, event ):
		selectedDataset = self.comboBox_selectModel.Value
		selectedMasterNetworkName = self.comboBox_selectNetwork.Value
		selectedScenarioName = self.comboBox_selectScenario.Value


		# Check whether user select needed items correctly
		message = ''
		if (selectedDataset == None or selectedDataset == ''):
			message = 'Select the model name in WamDam.'
		elif selectedMasterNetworkName == None or selectedMasterNetworkName == '':
			message = 'Select the MasterNetworkName.'
		elif selectedScenarioName == None or selectedScenarioName == '':
			message = 'Select the ScenarioName.'
		elif not ['xls', 'xlsx', 'xlsm', 'xls'].__contains__(self.path.split('.')[-1]):
			message = 'Please select a valid excel file.'

		if message != '':
			messageDlg = msg_somethigWrong(None, msg=message)
			messageDlg.ShowModal()
			return

		print 'this is done'

		try:
			exportTemplate = ExportTemplate(self.path)

			result = GetMetadata()
			organization_list = result.GetOrganizations(selectedDataset)
			exportTemplate.exportOrganizations(organization_list)

			people_list = result.GetPeople(selectedDataset)
			exportTemplate.exportPeople(people_list)

			source_list = result.GetSources(selectedDataset)
			exportTemplate.exportSources(source_list)

			# self.write2excel(source_list, 8, 3)
			method_list = result.GetMethods(selectedDataset)
			exportTemplate.exportMethods(method_list)

			###############################################
			result = GetInstances()

			names, network_data_result = result.getMasterNetwork(selectedDataset)
			exportTemplate.exportMasterNetwork(network_data_result)

			names, scenarios_data_result = result.getScenario(selectedDataset, selectedMasterNetworkName)
			exportTemplate.exportScenario(scenarios_data_result)

			names, nodes_data_result = result.getNodes(selectedDataset, selectedMasterNetworkName,selectedScenarioName)
			print "Count of Nodes data: {}".format(str(len(nodes_data_result)))
			exportTemplate.exportNodes(nodes_data_result)

			links_data_result = result.getLinkes(selectedDataset, selectedMasterNetworkName,selectedScenarioName)
			print "Count of links data: {}".format(str(len(links_data_result)))
			exportTemplate.exportLinkes(links_data_result)

			###############################################
			result = GetDataStructure()

			data_result = result.getObjecttypes(selectedDataset)
			exportTemplate.exportObjecttypes(data_result)

			attributes_result = result.getAttributes(selectedDataset)
			print "Count of attrs data: {}".format(str(len(attributes_result)))
			exportTemplate.exportAttributes(attributes_result)

			resources_result = result.getResourceType(selectedDataset)
			exportTemplate.exportResourcesType(resources_result)
			# self.write2excel(dataset_result, 10, 4)


			# Get all values of attributes for all the instances of the network

			###############################################
			result = GetAllValuesByScenario()

			result_list = result.GetAllTextFree(selectedDataset, selectedMasterNetworkName, selectedScenarioName)
			print "Count of TextFree data: {}".format(str(len(result_list)))
			exportTemplate.exportFreeText(result_list)

			result_list = result.GetAllNumericValues(selectedDataset, selectedMasterNetworkName, selectedScenarioName)
			print "Count of Numeric data: {}".format(str(len(result_list)))
			exportTemplate.exportNumericValue(result_list)

			result_list = result.GetAllCategoricalValues(selectedDataset, selectedMasterNetworkName, selectedScenarioName)
			print "Count of Categorical data: {}".format(str(len(result_list)))
			exportTemplate.exportCategoricalValues(result_list)

			result_list = result.GetAllSeasonalNumericValues(selectedDataset, selectedMasterNetworkName, selectedScenarioName)
			print "Count of SeasonalNumeric data: {}".format(str(len(result_list)))
			exportTemplate.exportSeasonal(result_list)

			result_list = result.GetAllTimeSeries(selectedDataset, selectedMasterNetworkName, selectedScenarioName)
			print "Count of Timeseries data: {}".format(str(len(result_list)))
			exportTemplate.exportTimeSeries(result_list)

			up_table_column_result, bottom_table_result = result.GetAllMultiAttributeSeries(selectedDataset, selectedMasterNetworkName, selectedScenarioName)
			exportTemplate.exportMulti(up_table_column_result, bottom_table_result)

			from Messages_forms.msg_successLoadDatabase import msg_successLoadDatabase
			instance = msg_successLoadDatabase(None)
			instance.m_staticText1.SetLabel("Success export excel file")
			instance.ShowModal()
		except Exception as e:
			messageDlg = msg_somethigWrong(None, msg=e.message)
			messageDlg.Show()
			# raise Exception(e.message)

	def write2excel(self, resutl_list, startRow, sheet):
		try:
			for row_id, row in enumerate(resutl_list):
				for col_id, cell in enumerate(row):
					sheet.cell(row=row_id + startRow, column=col_id + 1, value=cell)
			# if self.path.split('.')[-1] == 'xls':
			# 	# try:
			# 	# 	sheet = self.workbook.get_sheet(sheetNumber)
			# 	# except:
			# 	# 	# message = msg_somethigWrong(None, msg="Please select a valid Excel File")
			# 	# 	# message.ShowModal()
			# 	# 	raise Exception("Please select a valid Excel File")
			# 	for row_id, row in enumerate(resutl_list):
			# 		for col_id, cell in enumerate(row):
			# 			sheet.write(row_id + startRow - 1, col_id + 0, cell)
			# 	# self.workbook.save(self.path)
			# else:
			# 	''' Validate sheet in excel'''
			# 	try:
			# 		sheet = self.workbook.get_sheet_by_name(self.sheetNames[sheetNumber])
			# 	except:
			# 		# message = msg_somethigWrong(None, msg="Please select a valid Excel File")
			# 		# message.ShowModal()
			# 		# return
			# 		raise Exception("Please select a valid Excel File")
			# 	for row_id, row in enumerate(resutl_list):
			# 		for col_id, cell in enumerate(row):
			# 			sheet.cell(row=row_id + startRow, column=col_id + 1, value=cell)
            #
			# 	# self.save(self.path)

		except Exception as e:
			print e
			messageDlg = msg_somethigWrong(None, msg=e.message)
			messageDlg.Show()
			raise Exception(e.message)
	def btn_cancelOnButtonClick( self, event ):
		# TODO: Implement btn_cancelOnButtonClick
		self.Close()
	
	
