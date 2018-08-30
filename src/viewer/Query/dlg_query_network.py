"""Subclass of dlg_query_network, which is generated by wxFormBuilder."""

import wx
import viewer.WaMDaMWizard

from controller.wamdamAPI.GetDataStructure import GetDataStructure
from xlrd import open_workbook
# from xlutils.copy import copy
from viewer.Messages_forms.msg_somethigWrong import msg_somethigWrong
# This library is used here to write data to an excel file
from openpyxl import load_workbook
from controller.ConnectDB_ParseExcel import DB_Setup

# Implementing dlg_query_network
class dlg_query_network(viewer.WaMDaMWizard.dlg_query_network):
	def __init__( self, parent ):
		viewer.WaMDaMWizard.dlg_query_network.__init__(self, parent)

		self.path = ''
		try:
			if not DB_Setup().get_session():
				msg = "\n\nWarning: Please connect to sqlite first."
				raise Exception(msg)

			self.dataStructure = GetDataStructure()
			self.datasets = self.dataStructure.getResourceTypes()
			list_acromy = list()
			for row in self.datasets:
				list_acromy.append(row.ResourceTypeAcronym)
			if list_acromy.__len__() > 0:
				self.comboBox_selectModel.SetItems(list_acromy)
		except Exception as e:
			message = msg_somethigWrong(None, msg=e.message)
			message.ShowModal()
			self.Destroy()
	
	# Handlers for dlg_query_network events.
	def comboBox_selectModelOnCombobox( self, event ):
		# TODO: Implement comboBox_selectModelOnCombobox
		selectedDataset = self.comboBox_selectModel.Value
		result = GetDataStructure()
		gettedMasterNetworkNames, data = result.getMasterNetwork(selectedDataset)
		if gettedMasterNetworkNames.__len__() <= 0:
			return
		self.comboBox_selectNetwork.SetItems(gettedMasterNetworkNames)
	
	def comboBox_selectNetworkOnCombobox( self, event ):
		# TODO: Implement comboBox_selectNetworkOnCombobox
		selectedMasterNetworkName = self.comboBox_selectNetwork.Value
		result = GetDataStructure()
		gettedScenarioNames, data = result.getScenario(self.comboBox_selectModel.Value, selectedMasterNetworkName)
		if gettedScenarioNames.__len__() <= 0:
			return
		self.comboBox_selectScenario.SetItems(gettedScenarioNames)
		pass

	def comboBox_selectScenarioOnCombobox( self, event ):
		# TODO: Implement comboBox_selectScenarioOnCombobox
		pass

	def FilePicker_QueryNetworkOnFileChanged( self, event ):
		# TODO: Implement FilePicker_QueryNetworkOnFileChanged
		valid_extension = ['xls','xlsx']
		self.path = self.FilePicker_QueryNetwork.GetPath()
		if not (self.path.split('.')[-1] in valid_extension):
			message = msg_somethigWrong(None, msg="Please select a valid Excel File")
			message.Show()
			return
		print 'This is working just fine...'
		print self.path

	def btn_find_network_nodesLinksOnButtonClick( self, event ):
		# TODO: Implement btn_find_network_nodesLinksOnButtonClick
		print 'this is starting'
		selectedDataset = self.comboBox_selectModel.Value
		selectedMasterNetworkName = self.comboBox_selectNetwork.Value
		selectedScenarioName = self.comboBox_selectScenario.Value

		message = ''
		if (selectedDataset == None or selectedDataset == ''):
			message = 'Select the model name in WamDam.'
		elif selectedMasterNetworkName == None or selectedMasterNetworkName == '':
			message = 'Select the MasterNetworkName.'
		elif selectedScenarioName == None or selectedScenarioName == '':
			message = 'Select the ScenarioName.'
		elif not ['xls', 'xlsx', 'xlsm', 'xls'].__contains__(self.path.split('.')[-1]):
			message = 'And please select a valid excel file.'

		if message != '':
			messageDlg = msg_somethigWrong(None, msg=message)
			messageDlg.Show()
			raise Exception(message)

		''' get nework, scenario, nodes, links'''
		result = GetDataStructure()
		names, network_data_result = result.getMasterNetwork(selectedDataset)
		names, scenarios_data_result = result.getScenario(selectedDataset, selectedMasterNetworkName)
		names, nodes_data_result = result.getNodes(selectedDataset, selectedMasterNetworkName,selectedScenarioName)
		links_data_result = result.getLinkes(selectedDataset, selectedMasterNetworkName,selectedScenarioName)
		print '************8**'
		try:
			if self.path.split('.')[-1] == 'xls':
				wb = open_workbook(self.path)
				# workbook = copy(wb)
				workbook = wb
				try:
					sheet = workbook.get_sheet(6)
					sheetnodes = workbook.get_sheet(7)
					sheetlinks = workbook.get_sheet(8)
				except:
					message = msg_somethigWrong(None, msg="Please select a valid Excel File")
					message.Show()
					return

				for row_id, row in enumerate(network_data_result):
					for col_id, cell in enumerate(row):
						sheet.write(row_id + 9, col_id + 0, cell)

				for row_id, row in enumerate(scenarios_data_result):
					for col_id, cell in enumerate(row):
						sheet.write(row_id + 17, col_id + 0, cell)


				for row_id, row in enumerate(nodes_data_result):
					for col_id, cell in enumerate(row):
						sheetnodes.write(row_id + 10, col_id + 0, cell)

				for row_id, row in enumerate(links_data_result):
					for col_id, cell in enumerate(row):
						sheetlinks.write(row_id + 10, col_id + 0, cell)

				workbook.save(self.path)
			else:
				book2 = load_workbook(self.path)
				try:
					sheet = book2.get_sheet_by_name("3.1_Networks&Scenarios")
					nodes_sheet = book2.get_sheet_by_name("3.2_Nodes")
					links_sheet = book2.get_sheet_by_name("3.3_Links")
				except:
					message = msg_somethigWrong(None, msg="Please select a valid Excel File")
					message.Show()
					return
					# raise Exception('Output Sheet {} not found in Excel File \n\n Please select valid Excel File'.format("3.1_Networks&Scenarios"))

				for row_id, row in enumerate(network_data_result):
					for col_id, cell in enumerate(row):
						sheet.cell(row=row_id + 10, column=col_id + 1, value=cell)

				for row_id, row in enumerate(scenarios_data_result):
					for col_id, cell in enumerate(row):
						sheet.cell(row=row_id + 20, column=col_id + 1, value=cell)

				for row_id, row in enumerate(nodes_data_result):
					for col_id, cell in enumerate(row):
						nodes_sheet.cell(row=row_id + 11, column=col_id + 1, value=cell)

				for row_id, row in enumerate(links_data_result):
					for col_id, cell in enumerate(row):
						links_sheet.cell(row=row_id + 11, column=col_id + 1, value=cell)
				book2.save(self.path)
			from Messages_forms.msg_successLoadDatabase import msg_successLoadDatabase
			instance = msg_successLoadDatabase(None)
			instance.m_staticText1.SetLabel("Success: exported excel file")
			instance.ShowModal()
		except Exception as e:
			print e
			messageDlg = msg_somethigWrong(None, msg=e.message)
			messageDlg.Show()
			raise Exception(e.message)

		print 'Great so far....'
	
	def btn_cancelOnButtonClick( self, event ):
		# TODO: Implement btn_cancelOnButtonClick
		self.Close()
	
	
