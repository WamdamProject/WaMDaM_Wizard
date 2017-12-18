"""Subclass of dlg_SearchNodesLinks, which is generated by wxFormBuilder."""

import wx
import WaMDaMWizard
from controller.wamdamAPI.GetNodeLinks import GetNodeLinks
from xlrd import open_workbook
# from xlutils.copy import copy
from Messages_forms.msg_somethigWrong import msg_somethigWrong
# This library is used here to write data to an excel file
from openpyxl import load_workbook

# Implementing dlg_SearchNodesLinks
class dlg_SearchNodesLinks( WaMDaMWizard.dlg_SearchNodesLinks ):
	def __init__( self, parent ):
		WaMDaMWizard.dlg_SearchNodesLinks.__init__( self, parent )
		self.path = ''
		# self.textCtrl_X_min.Value = u'-112.4424'
		# self.textCtrl_X_Max.Value = u'-110.65833'
		# self.textCtrl_y_min.Value = u'41.00'
		# self.textCtrl_y_max.Value = u'42.700'
		try:
			''' Get Controlled object type'''
			self.getNodeLinks = GetNodeLinks()
			list_acromy = self.getNodeLinks.GetCV_ObjectType()
			if list_acromy.__len__() > 0:
				self.comboBox_selectObjectType.SetItems(list_acromy)
		except Exception as e:
			message = msg_somethigWrong(None, msg=e.message)
			message.Show()
			self.Destroy()
	# Handlers for dlg_SearchNodesLinks events.
	def checkBox_NodesOnCheckBox( self, event ):
		# TODO: Implement checkBox_NodesOnCheckBox
		pass
	
	def comboBox_selectObjectTypeOnCombobox( self, event ):
		# TODO: Implement comboBox_selectObjectTypeOnCombobox
		pass
	
	def textCtrl_X_minOnText( self, event ):
		# TODO: Implement textCtrl_X_minOnText
		pass
	
	def textCtrl_X_MaxOnText( self, event ):
		# TODO: Implement textCtrl_X_MaxOnText
		pass
	
	def textCtrl_y_minOnText( self, event ):
		# TODO: Implement textCtrl_y_minOnText
		pass
	
	def textCtrl_y_maxOnText( self, event ):
		# TODO: Implement textCtrl_y_maxOnText
		pass
	
	def FilePicker_searchNodesLinksOnFileChanged( self, event ):
		# TODO: Implement FilePicker_searchNodesLinksOnFileChanged
		valid_extension = ['xls','xlsx']
		self.path = self.FilePicker_searchNodesLinks.GetPath()
		if not (self.path.split('.')[-1] in valid_extension):
			message = msg_somethigWrong(None, msg="Please select a valid Excel File")
			message.ShowModal()
			return
		print 'This is working just fine...'
		print self.path
	
	def btn_Search_nodesLinksOnButtonClick( self, event ):
		# TODO: Implement btn_Search_nodesLinksOnButtonClick

		'''Check whether longitude is between -180 ~ 180.'''
		try:
			if (float(self.textCtrl_X_min.Value) < -180 or float(self.textCtrl_X_Max.Value) >180 ):
				message = msg_somethigWrong(None, msg="Please input longitude between -180 ~ 180")
				message.ShowModal()
				return
		except:
			message = msg_somethigWrong(None, msg="Please input float type in Minimum (East) or Maximum (West).")
			message.ShowModal()
			return
		'''Check whether latitude is between -180 ~ 180.'''
		try:
			if (float(self.textCtrl_y_min.Value) < -90 or float(self.textCtrl_y_max.Value) >90 ):
				message = msg_somethigWrong(None, msg="Please input latitude between -90 ~ 90")
				message.ShowModal()
				return
		except:
			message = msg_somethigWrong(None, msg="Please input float type in Minimum (South) or Maximum (North).")
			message.ShowModal()
			return
		'''Check whether Controlled Object Type is selected and selected excel file is valid.'''
		message = ''
		if self.comboBox_selectObjectType.Value == None or self.comboBox_selectObjectType.Value == '':
			message = 'Please select Controlled Object Type!!!'
		elif not ['xlsx', 'xlsm', 'xls'].__contains__(self.path.split('.')[-1]):
			message = 'please select a valid excel file.'

		if message != '':
			messageDlg = msg_somethigWrong(None, msg=message)
			messageDlg.Show()
			return

		print '************8**'
		try:
			xmin = float(self.textCtrl_X_min.Value)
			ymin = float(self.textCtrl_y_min.Value)
			xmax = float(self.textCtrl_X_Max.Value)
			ymax = float(self.textCtrl_y_max.Value)

			''' Get MasterNetwork, Scenarios, Nodes, Links'''
			networks = self.getNodeLinks.getMasterNetwork(self.comboBox_selectObjectType.Value, xmin, ymin, xmax, ymax)
			if len(networks) < 1:
				messageDlg = msg_somethigWrong(None, msg="Nothing found that meets this query result")
				messageDlg.Show()
				return
			scenarios = self.getNodeLinks.GetScenaroisResult(self.comboBox_selectObjectType.Value, xmin, ymin, xmax, ymax)
			nodes = self.getNodeLinks.GetNodesResult(self.comboBox_selectObjectType.Value, xmin, ymin, xmax, ymax)
			links = self.getNodeLinks.GetLinksResult(self.comboBox_selectObjectType.Value, xmin, ymin, xmax, ymax)

			if self.path.split('.')[-1] == 'xls':
				wb = open_workbook(self.path)
				# workbook = copy(wb)
				workbook = wb
				try:
					sheet = workbook.get_sheet(6)
					sheetnodes = workbook.get_sheet(7)
					sheetlinks = workbook.get_sheet(8)
				except:
					message = msg_somethigWrong(None, msg="Please select a valid Excel File")
					message.Show()
					return

				for row_id, row in enumerate(networks):
					for col_id, cell in enumerate(row):
						sheet.write(row_id + 9, col_id + 0, cell)

				for row_id, row in enumerate(scenarios):
					for col_id, cell in enumerate(row):
						sheet.write(row_id + 17, col_id + 0, cell)

				for row_id, row in enumerate(nodes):
					for col_id, cell in enumerate(row):
						sheetnodes.write(row_id + 10, col_id + 0, cell)

				for row_id, row in enumerate(links):
					for col_id, cell in enumerate(row):
						sheetlinks.write(row_id + 10, col_id + 0, cell)

				workbook.save(self.path)
			else:
				book2 = load_workbook(self.path)
				try:
					sheet = book2.get_sheet_by_name("3.1_Networks&Scenarios")
					nodes_sheet = book2.get_sheet_by_name("3.2_Nodes")
					links_sheet = book2.get_sheet_by_name("3.3_Links")
				except:
					message = msg_somethigWrong(None, msg="Please select a valid Excel File")
					message.Show()
					return
					# raise Exception('Output Sheet {} not found in Excel File \n\n Please select valid Excel File'.format("3.1_Networks&Scenarios"))

				for row_id, row in enumerate(networks):
					for col_id, cell in enumerate(row):
						sheet.cell(row=row_id + 10, column=col_id + 1, value=cell)

				for row_id, row in enumerate(scenarios):
					for col_id, cell in enumerate(row):
						sheet.cell(row=row_id + 20, column=col_id + 1, value=cell)

				for row_id, row in enumerate(nodes):
					for col_id, cell in enumerate(row):
						nodes_sheet.cell(row=row_id + 11, column=col_id + 1, value=cell)

				for row_id, row in enumerate(links):
					for col_id, cell in enumerate(row):
						links_sheet.cell(row=row_id + 11, column=col_id + 1, value=cell)
				book2.save(self.path)
			from Messages_forms.msg_successLoadDatabase import msg_successLoadDatabase
			instance = msg_successLoadDatabase(None)
			instance.m_staticText1.SetLabel("Success export excel file")
			instance.ShowModal()
		except Exception as e:
			print e
			messageDlg = msg_somethigWrong(None, msg=e.message)
			messageDlg.Show()
			raise Exception(e.message)

		print 'Great so far....'
	def btn_cancelOnButtonClick( self, event ):
		# TODO: Implement btn_cancelOnButtonClick
		self.Close()
