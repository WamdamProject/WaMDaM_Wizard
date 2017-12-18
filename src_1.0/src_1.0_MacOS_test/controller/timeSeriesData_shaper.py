
"""
    The timeSeriesData_shaper.py is a python script use to reshape
    TimeSeriesData and write the output result in the appropriate
    excel sheets. 
    It has three functions:
    
    1. 
    2. 
    3. 


"""

# The xlrd library is used for only reading data in excel files. its is quite fast at that
import xlrd as excel

# This library is used here to write data to an excel file
from openpyxl import load_workbook
import define

# This library is used to get options from user if script is run from terminal
from optparse import OptionParser


# ****************************************************************************************************************** #
#                                                                                                                    #
#                                         Script Begins here                                                         #
#                                                                                                                    #
# ****************************************************************************************************************** #


class TimeSeriesDataShaperClass():
    """
        This is a helper class for task1
    """
    month, sheet, rows = ([], [], [])

    def __init__(self):
        pass

    def set_params(self, sheet):
        """
        This function allows task1() to
        setup important parameters for
        successful completion of its task.

        :param sheet: it takes one parameter
        (sheet name) and sets up the sheet
        name for writing data, the rows of
        the sheets and the months.

        :return: None
        """

        """ get sheet header's information"""
        self.sheet = sheet
        if self.sheet.name == "TimeSeries_month1_input":
            self.temp = self.get_sheet_rows()
            self.fieldNames = self.temp[0][:5]
            self.fieldValues = self.temp[1][:5]
            self.rows = self.temp[6:]
            self.fieldNames1 = self.rows[0][:2]
            self.month = self.months()
        elif self.sheet.name == "TimeSeries_month2_input":
            self.temp = self.get_sheet_rows()
            self.fieldNames = self.temp[0][:5]
            self.fieldValues = self.temp[1][:5]
            self.rows = self.temp[4:]
            self.month = self.rows.pop(0)[1:]
            self.year = self.rows.pop(0)[1:]
            self.fieldNames1 = self.rows.pop(0)[0]
        elif self.sheet.name == "TimeSeries_daily_input3":
            self.temp = self.get_sheet_rows()
            self.fieldNames = self.temp[0][:5]
            self.fieldValues = self.temp[1][:5]
            self.days = self.temp[5][3:]
            self.rows = self.temp[7:]
            self.fieldNames1 = self.rows[0][:3]
            self.month = self.rows.pop(0)[3:]
        else:

            raise Exception("Error\nCould not find " + self.sheet.name + ".\nPlease input again!!!")

        self.rowCount = self.rows.__len__()


        print(self.fieldNames)
        print(self.fieldNames1)
        print(self.month)

    def get_excel_frame(self, filename):
        """
        Takes an excel workbook and return a
        pointer to the workbook which permits
        access to the files data

        :param filename: Name of workbook

        :return: pointer to workbook
        """
        return excel.open_workbook(filename)

    def get_sheet_rows(self):
        """
        function gets all the rows from sheet
        in the workbook

        :return: rows in the specified sheet
        """
        return [row for row in self.sheet.get_rows()]

    def create_matrics(self):
        """
        This function creates two matrices, one is
        the instance_matrix(properties) and the other
        is the timeseries_matrix(from timeseries columns)

        :return: two matrices (properties) and (timeseries cols)
        """

        """
        real data first column number set"""

        n = 0
        if self.sheet.name == "TimeSeries_month1_input":
            n = 2
        elif self.sheet.name == "TimeSeries_month2_input":
            n = 1
        elif self.sheet.name == "TimeSeries_daily_input3":
            n = 3
        else:
            raise Exception("Error\nCould not find " + self.sheet.name + ".\nPlease input again!!!")



        instance_matrix, timeseries_matrix = (list(), list())
        for i in range(self.rows.__len__()):
            row = self.rows[i]
            temp = []
            m = 0

            """Here get instance, year and day matrix."""

            for cell in row[:n]:
                if str(cell.value) == "":
                    raise Exception("Error!\nin the sheet " + self.sheet.name + ",\nCell that field name is '"
                                    + self.fieldNames1[m].value + "' and row is " + str(i + 1) + " is empty!")
                temp.append(str(cell.value))
                m += 1
            instance_matrix.append(temp)
            # instance_matrix.append([str(cell.value) for cell in row[:n]])
            temp = []
            m = 0

            """
                Here get time series matrix.
                first, check whether value of cell is empty or mot.
                next, if no empty, add cell value to time series matrix.
            """

            for cell in row[n:]:
                if str(cell.value) == "":
                    if self.sheet.name == "TimeSeries_month2_input":
                        raise Exception("Error!\nin the sheet " + self.sheet.name + ",\nCell that field is '"
                                    + self.month[m].value + "/" + str(self.year[m].value).split('.')[0] + "' and row is " + str(i + 1) + " is empty!")
                    else:
                        if self.sheet.name == "TimeSeries_daily_input3":
                            if self.month[m].value == 'February' and [28, 29, 30].__contains__(i):
                                m += 1
                                temp.append('')
                                continue
                            elif ['November', 'April', 'June', 'September'].__contains__(self.month[m].value) and i == 30:
                                m += 1
                                temp.append('')
                                continue
                        raise Exception("Error!\nin the sheet " + self.sheet.name + ",\nCell that field name is '"
                                        + self.month[m].value + "' and row is " + str(i + 1) + " is empty!")
                temp.append(str(cell.value))
                m += 1
            timeseries_matrix.append(temp)
        return instance_matrix, timeseries_matrix

    def cross_matrics(self):
        """
        This function does all the heavy lifting. it takes each row in
        instance_matrix and crosses it with the corresponding column
        in timeseries_matix. (This is the abstraction behind the below code)

        :return: crossed matrix
        """
        """ get header's information(fVal) """
        all = list()
        fVal = list()
        for cell in self.fieldValues:
            fVal.append(cell.value)

        """get data(all) to output from instance amd time series matrix.
           data varies from table to table. """

        instance_matrix, timeseries_matrix = self.create_matrics()
        if self.sheet.name == "TimeSeries_month1_input":
            for instance, timeseries in zip(instance_matrix, timeseries_matrix):
                for index, value in enumerate(timeseries):
                    if len(fVal) > 5:
                        fVal.pop(1)
                    fVal.insert(1, instance[0])
                    temp = fVal[:]
                    monthVal = (index + 10) % 12
                    if monthVal == 0:
                        monthVal = 12
                    temp.append(str(monthVal) + '/1/' + instance[1].split('.')[0])
                    temp.append(value)
                    all.append(temp)
            return all
        elif self.sheet.name == "TimeSeries_month2_input":
            for instance, timeseries in zip(instance_matrix, timeseries_matrix):
                for index, value in enumerate(timeseries):
                    if len(fVal) > 5:
                        fVal.pop(1)
                    fVal.insert(1, instance[0])
                    temp = fVal[:]
                    monthVal = (index + 1) % 12
                    if monthVal == 0:
                        monthVal = 12
                    temp.append(str(monthVal) + '/1/' + str(self.year[index].value).split('.')[0])
                    temp.append(value)
                    all.append(temp)
            return all
        elif self.sheet.name == "TimeSeries_daily_input3":
            for index in range(12):
                for instance, timeseries in zip(instance_matrix, timeseries_matrix):
                    if len(fVal) > 5:
                        fVal.pop(1)
                    fVal.insert(1, instance[0])
                    temp = fVal[:]
                    monthVal = (index + 10) % 12
                    if monthVal == 0:
                        monthVal = 12
                    temp.append(str(monthVal) + '/' + str(instance[2].split('.')[0]) + '/' + str(str(instance[1])).split('.')[0])
                    temp.append(timeseries[index])
                    all.append(temp)
            return all

    def months(self):
        """
        Gets all months of each timeseries column
        :return:
        """
        return self.rows.pop(0)[2:]

"""This function make timeSeries sheet """
def WriteToExcelFile(workbook,
          input_sheet='TimeSeries_month1_input', output_sheet='TimeSeries_month1_output'):

    """
    :param workbook: excel file
    :param input_sheet: convert sheet
    :param output_sheet: conveted sheet
    :return: none
    """

    instance = TimeSeriesDataShaperClass()
    book = instance.get_excel_frame(workbook)
    try:
        sheet = book.sheet_by_name(input_sheet)
    except:
        raise Exception('Input Sheet {} not found in {} \n Please select valid Excel file'.format(input_sheet, workbook))
    instance.set_params(sheet)

    final_data = instance.cross_matrics()

    book2 = load_workbook(workbook)
    try:
        sheet = book2.get_sheet_by_name(output_sheet)
    except:
        raise Exception('Output Sheet {} not found in {} \n Please select valid Excel File'.format(output_sheet, workbook))

    for row_id, row in enumerate(final_data):
        for col_id, cell in enumerate(row):
            sheet.cell(row=row_id + 10, column=col_id + 1, value=cell)
    book2.save(workbook)

def TimeSeriesDataShaper(workbook):

    """
    read 3 sheets at one time
    """
    WriteToExcelFile(workbook, 'TimeSeries_month1_input', 'TimeSeries_month1_ouput')
    WriteToExcelFile(workbook, 'TimeSeries_month2_input', 'TimeSeries_month2_output')
    WriteToExcelFile(workbook, 'TimeSeries_daily_input3', 'TimeSeries_daily_output')




