

"""
    The seasonalData_shaper.py is a python script use to reshape
    SeasonalData and write the output result in the appropriate
    excel sheets. it is made up of:

    - SeasonalDataShaper():  This is the function which restructures the Seasonal
                data by doing some cross multiplications of the Seasonal
                properties and the SeasonName Columns.
                NB!! if format of file changes, script will fail to work

    - SeasonalDataShaper:  This is a class which helps in achieving the goal of
                            SeasonalDataShaper(). it provides some basic functions accessible
                            to SeasonalDataShaper() so as to render the function less combersome.
"""

# The xlrd library is used for only reading data in excel files. its is quite fast at that
import xlrd as excel

# This library is used here to write data to an excel file
from openpyxl import load_workbook

# This library is used to get options from user if script is run from terminal
from optparse import OptionParser


# ****************************************************************************************************************** #
#                                                                                                                    #
#                                         Script Begins here                                                         #
#                                                                                                                    #
# ****************************************************************************************************************** #


class SeasonalDataShaper():
    """
        This is a helper class for task1
    """
    month, sheet, rows = ([], [], [])

    def __init__(self):
        pass

    def set_params(self, sheet):
        """
        This function allows SeasonalDataShaper() to
        setup important parameters for
        successful completion of its task.

        :param sheet: it takes one parameter
        (sheet name) and sets up the sheet
        name for writing data, the rows of
        the sheets and the months.

        :return: None
        """
        self.sheet = sheet
        self.temp = self.get_sheet_rows()
        self.rows = self.temp[2:]
        self.month = self.months()
        self.cv = self.seasonCV()
        print self.cv
        print self.month

    def get_excel_frame(self, filename):
        """
        Takes an excel workbook and return a
        pointer to the workbook which permits
        access to the files data

        :param filename: Name of workbook

        :return: pointer to workbook
        """
        return excel.open_workbook(filename)

    def get_sheet_rows(self):
        """
        function gets all the rows from sheet
        in the workbook

        :return: rows in the specified sheet
        """
        return [row for row in self.sheet.get_rows()][1:]

    def create_matrics(self):
        """
        This function creates two matrices, one is
        the instance_matrix(properties) and the other
        is the season_matrix(from SeasonName columns)

        :return: two matrices (properties) and (seasonName cols)
        """
        instance_matrix, season_matrix = (list(), list())
        for row in self.rows:
            instance_matrix.append([str(cell.value) for cell in row[:6]])
            season_matrix.append([str(cell.value) for cell in row[6:]])
        return instance_matrix, season_matrix

    def cross_matrics(self):
        """
        This function does all the heavy lifting. it takes each row in
        instance_matrix and crosses it with the corresponding column
        in season_matix. (This is the abstraction behind the below code)

        :return: crossed matrix
        """
        all = list()
        instance_matrix, season_matrix = self.create_matrics()
        for instance, season in zip(instance_matrix, season_matrix):
            for index, value in enumerate(season):
                temp = instance[:]
                temp.append(str(self.month[index].value))
                temp.append(str(self.cv[index].value))
                temp.append(value)
                all.append(temp)
        return all



    def months(self):
        """
        Gets all months of each seasonName column
        :return:
        """
        return self.rows.pop(0)[6:]

    def seasonCV(self):
        """
        gets all the season name cv
        :return:
        """
        return self.temp[0][6:]

def SeasonalDataShaperFunc(workbook,
          input_sheet='SeasonalParameter_input', output_sheet='SeasonalNumericValues_input'):
    """
        This function reshapes the Seasonal Data (SeasonalTemplate_input)
         input sheet with the help of its class and write parsed result
         to the output_sheet (SeasonalParameter)
        :param workbook: variable holds the Seasonal excel file
        :param input_sheet: hold the seasonal input sheet
        :param output_sheet: holds the output sheet to write shaped data
        :return: None
    """
    instance = SeasonalDataShaper()
    book = instance.get_excel_frame(workbook)
    try:
        sheet = book.sheet_by_name(input_sheet)
    except:
        raise Exception('Input sheet {} not found in Excel File \n Please select valid Excel file'.format(input_sheet))
    instance.set_params(sheet)

    final_data = instance.cross_matrics()

    book2 = load_workbook(workbook)
    try:
        sheet = book2.get_sheet_by_name(output_sheet)
    except:
        raise Exception('Output sheet {} not found in Excel File \n Please select valid Excel File'.format(input_sheet))

    for row_id, row in enumerate(final_data):
        for col_id, cell in enumerate(row):
            sheet.cell(row=row_id + 10, column=col_id + 1, value=cell)
    book2.save(workbook)


