
from ..ConnectDB_ParseExcel import DB_Setup
from ..ConnectDB_ParseExcel import SqlAlchemy as sq
from openpyxl import load_workbook
from xlrd import open_workbook
# from xlutils.copy import copy

'''
This class is used to get data values.
'''
class GetDataValues(object):
    def __init__(self):
        self.setup = DB_Setup()
        self.session = self.setup.get_session()
    def GetCV_ObjectType(self):
        '''
        This method is used to get all names of CV_ObjectType.
        :return: list of CV_ObjectType's name
        '''
        result = self.session.query(sq.CV_ObjectType.Name).all()
        # complete_result = list()
        nameResult = list()
        for row in result:
            isExisting = False
            for name in nameResult:
                if name == row.Name:
                    isExisting = True
                    break
            if not isExisting:
                nameResult.append(row.Name)
        return nameResult

    def getAttribute(self):
        '''
        This method is used to get all names of CV_AttributeName.
        :return: list of CV_AttributeName's name
        '''
        result = self.session.query(sq.CV_AttributeName.Name).all()
        # complete_result = list()
        nameResult = list()
        for row in result:
            isExisting = False
            for name in nameResult:
                if name == row.Name:
                    isExisting = True
                    break
            if not isExisting:
                nameResult.append(row.Name)
        return nameResult

    def getCVinstanceName(self):
        '''
        This method is used to get all names of CV_InstanceName.
        :return: list of CV_InstanceName's name
        '''
        result = self.session.query(sq.CV_InstanceName.Name).all()
        # complete_result = list()
        nameResult = list()
        for row in result:
            isExisting = False
            for name in nameResult:
                if name == row.Name:
                    isExisting = True
                    break
            if not isExisting:
                nameResult.append(row.Name)
        return nameResult
    def check_validation(self):
        '''
        This method is used to check validations of bellow tables within db.
        :return: None
        '''
        table_list = ["DualValues", "DescriptorValues", "NumericValues", "CV_ElectronicFormat", "SeasonalParameters", "TimeSeries",
                      "MultiAttributeSeriesValues"]
        for table_name in table_list:
            recordCountResult = self.session.execute('SELECT COUNT(*) FROM {};'.format(table_name))
            i = 0
            for n in recordCountResult:
                i = int(n[0])
            if i == 0:
                raise Exception('Warning!\n{} table is empty. Please fill this table.'.format(table_name))

    def exportDualValuesSheet(self, selectedType, selectedAttribute, selectedInstance, excelPath):
        '''
        This method is used to get data making DualValues_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            # result = self.session.query(sq.Datasets.DatasetAcronym, sq.Attributes.AttributeName, sq.Attributes.UnitNameCV,
            #                 sq.Instances.InstanceName, sq.MasterNetworks.MasterNetworkName,
            #                 sq.Scenarios.ScenarioName, sq.Sources.SourceName, sq.Methods.MethodName,
            #                 sq.DualValues.DualValuesValueMeaningCV)\
            #     .join(sq.ObjectTypes, sq.ObjectTypes.DatasetID == sq.Datasets.DatasetID)\
            #     .join(sq.Attributes, sq.Attributes.ObjectTypeID == sq.ObjectTypes.ObjectTypeID)\
            #     .join(sq.Mapping, sq.Mapping.AttributeID == sq.Attributes.AttributeID)\
            #     .join(sq.DataValuesMapper, sq.DataValuesMapper.DataValuesMapperID == sq.Mapping.DataValuesMapperID)\
            #     .join(sq.ScenarioMapping, sq.ScenarioMapping.MappingID == sq.Mapping.MappingID)\
            #     .join(sq.Scenarios, sq.Scenarios.ScenarioID == sq.ScenarioMapping.ScenarioID)\
            #     .join(sq.MasterNetworks, sq.MasterNetworks.MasterNetworkID == sq.Scenarios.MasterNetworkID)\
            #     .join(sq.Methods, sq.Methods.MethodID == sq.Mapping.MethodID)\
            #     .join(sq.Sources, sq.Sources.SourceID == sq.Mapping.SourceID)\
            #     .join(sq.Instances, sq.Instances.InstanceID == sq.Mapping.InstanceID)\
            #     .join(sq.DualValues, sq.DualValues.DataValuesMapperID == sq.DataValuesMapper.DataValuesMapperID)\
            #     .join(sq.CV_DualValueMeaning, sq.CV_DualValueMeaning.Name == sq.DualValues.DualValuesValueMeaningCV)\
            #     .filter(sq.Attributes.AttributeDataTypeCV=='DualValues')\
            #     .all()

            sql = 'SELECT AttributeName, SourceName, InstanceName,MasterNetworkName,ScenarioName,MethodName,dualvaluemeaningCV ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "DualValues" ON "DualValues"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'LEFT JOIN "CV_DualValueMeaning" ON "CV_DualValueMeaning"."Name"= "DualValues"."dualvaluemeaningCV" '\
                'WHERE "AttributeDataTypeCV"="DualValues" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                .format( selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.dualvaluemeaningCV])
            self.isMatching_query(complete_result, "DualValues")
            self.write2excel(complete_result, '4_DualValues', 9, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Error occurred in reading Data Structure.\n' + e.message)

    def exportTextConrolledSheet(self, selectedType, selectedAttribute, selectedInstance, excelPath):
        '''
        This method is used to get data making DescriptorValues_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT Attributes.AttributeName, ObjectType, SourceName, InstanceName,MasterNetworkName,' \
                  'ScenarioName,MethodName,descriptorvalueCV ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "DescriptorValues" ON "DescriptorValues"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'LEFT JOIN "CV_DescriptorValues" ON "CV_DescriptorValues"."Name"= "DescriptorValues"."descriptorvalueCV" '\
                'WHERE "AttributeDataTypeCV"="DescriptorValues" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.descriptorvalueCV])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "DescriptorValues")
                self.write2excel(complete_result, '4_DescriptorValues', 13, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)

    def exportNumericValuesheet(self, selectedType, selectedAttribute, selectedInstance, excelPath):
        '''
        This method is used to get data making NumericValues_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT "Datasets"."DatasetName", ObjectType,AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                  'ScenarioName,MethodName, NumericValue ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "NumericValues" ON "NumericValues"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'WHERE "AttributeDataTypeCV"="Parameter" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.NumericValue])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "Parameter")
                self.write2excel(complete_result, '4_Parameter', 10, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportElectronicFilesSheet(self, selectedType, selectedAttribute, selectedInstance, excelPath):
        '''
        This method is used to get data making ElectronicFiles_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                  'ScenarioName,MethodName,FileName, ElectronicFileFormatCV, "File"."Description" ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "File" ON "File"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'Left JOIN "CV_ElectronicFormat" ON "CV_ElectronicFormat"."Name"="File"."ElectronicFileFormatCV" '\
                'WHERE "AttributeDataTypeCV"="File" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.FileName, row.ElectronicFileFormatCV, excelPath, row.Description])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "ElectronicFiles")
                self.write2excel(complete_result, '4_ElectronicFiles', 14, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportSeasonalSheet(self, selectedType, selectedAttribute, selectedInstance, excelPath):
        '''
        This method is used to get data making SeasonalNumericValues_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                  'ScenarioName,MethodName,SeasonName, SeasonValue, SeasonNameCV ' \
                'FROM "Attributes" '\
                'Left JOIN "ObjectTypes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "SeasonalParameters" ON "SeasonalParameters"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'WHERE "AttributeDataTypeCV"="SeasonalParameter" AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.SeasonName, row.SeasonNameCV, row.SeasonValue])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "SeasonalParameter")
                self.write2excel(complete_result, '4_SeasonaNumericValues', 11, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportTimeSeriesSheet(self, selectedType, selectedAttribute, selectedInstance, excelPath):
        '''
        This method is used to get data making TimeSeries_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT DatasetName ObjectType, AttributeName, SourceName, InstanceName,WaterOrCalendarYear,' \
                  'ScenarioName,MethodName,AggregationStatisticCV, AggregationInterval, IntervalTimeUnitCV,' \
                  'IsRegular, NoDataValue, "TimeSeries"."Description" ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "TimeSeries" ON "TimeSeries"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'WHERE AttributeName!="ObjectInstances"  AND AttributeDataTypeCV="TimeSeries" ' \
                  ' AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.WaterOrCalendarYear, row.AggregationStatisticCV, row.AggregationInterval,
                                        row.IntervalTimeUnit, row.IsRegular, row.NoDataValue, row.Description])

            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "TimeSeries")
                self.write2excel(complete_result, '4_TimeSeries', 15, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportTextFreeSheet(self, selectedType, selectedAttribute, selectedInstance, excelPath):
        '''
        This method is used to get data making TextFreeSheet.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT DatasetName ObjectType, AttributeName, SourceName, InstanceName,TextFreeValue,' \
                  'ScenarioName,MethodName ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "TextFree" ON "TextFree"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'WHERE AttributeDataTypeCV="TextFree" ' \
                  ' AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "AttributeNameCV" = "{}"'\
                .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,row.TextFreeValue])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "TextFree")
                self.write2excel(complete_result, '4_TextFree', 12, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportMultiSheet(self, selectedType, selectedAttribute, selectedInstance, excelPath):
        '''
        This method is used to get data making MultiAttributeSeries_table.
        :param selectedType: selected Object Type
        :param selectedAttribute: controlled Attribute
        :param selectedInstance: controlled Instance Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT "ObjectTypes"."ObjectType", "Attributes"."AttributeName", SourceName, InstanceName,' \
                  'ScenarioName,MethodName, "AttributesColumns"."AttributeName" AS "ColumName", "AttributesColumns"."UnitNameCV" AS "ColUnitName", "Value","ValueOrder" '\
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'Left JOIN "MultiAttributeSeries" ON "MultiAttributeSeries"."DataValuesMapperID"="DataValuesMapper"."DataValuesMapperID" '\
                'Left JOIN "DataValuesMapper" As "DataValuesMapperColumn" ON "DataValuesMapperColumn"."DataValuesMapperID"="MultiAttributeSeries"."AttriNameID" '\
                'Left JOIN "Mapping" As "MappingColumns" ON "MappingColumns"."DataValuesMapperID"="DataValuesMapperColumn"."DataValuesMapperID" '\
                'Left JOIN "Attributes" AS "AttributesColumns" ON "AttributesColumns"."AttributeID"="MappingColumns"."AttributeID" '\
                'Left JOIN "MultiAttributeSeriesValues" ON "MultiAttributeSeriesValues"."MultiAttributeSeriesID"="MultiAttributeSeries"."MultiAttributeSeriesID" '\
                'WHERE Attributes.AttributeDataTypeCV="MultiAttributeSeries" '\
                'AND "ObjectTypeCV" = "{}" AND "InstanceNameCV" = "{}" AND "Attributes"."AttributeNameCV" = "{}"'\
                'Order By ScenarioName, AttributeName,ValueOrder asc '\
                .format(selectedType, selectedInstance, selectedAttribute)

            result = self.session.execute(sql)

            '''Down Table(MultiVariableSeries_table Table) write'''
            complete_result = list()
            strAtrributName = ''
            valueOrder = None
            columnName = ''
            tempColumn = {}
            sourceName = ''
            i = 0
            currentrow = 0
            setNumber = 0
            for row in result:
                if row.ColumName == None or row.ColumName == "":
                    continue
                if strAtrributName != row.AttributeName:
                    strAtrributName = row.AttributeName
                    tempColumn[row.AttributeName] = []
                    tempColumn[row.AttributeName].append(row.ColumName)
                    columnName = row.ColumName
                if sourceName != row.ScenarioName:
                    sourceName = row.ScenarioName
                    setNumber = i
                    currentrow = 0
                if columnName != row.ColumName:
                    columnName = row.ColumName
                    currentrow = 0

                if row.ColumName in tempColumn[row.AttributeName]:
                    index = tempColumn[row.AttributeName].index(row.ColumName)
                    if index == 0:
                        complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName, row.AttributeName, row.SourceName,
                                        row.MethodName, row.Value])
                        i += 1
                    else:
                        complete_result[setNumber + currentrow].append(row.Value)
                        currentrow += 1
                else:
                    currentrow = 0
                    tempColumn[row.AttributeName].append(row.ColumName)
                    index = tempColumn[row.AttributeName].index(row.ColumName)
                    if index == 0:
                        complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName, row.AttributeName, row.SourceName,
                                        row.MethodName, row.Value])
                        i += 1
                    else:
                        complete_result[setNumber + currentrow].append(row.Value)
                        currentrow += 1
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "MultiVariableSeries")
                self.write2excel(complete_result, '4_MultiVariableSeries', 17, 15, excelPath)

            '''Up Table(AttributeName_column Table) write'''
            column_result = list()
            for key, columnItems in tempColumn.items():
                temList = list()
                temList.append(key)
                for item in columnItems:
                    temList.append(item)
                column_result.append(temList)
            if complete_result.__len__() > 0:
                self.write2excel(column_result, '4_MultiAttributeSeries', 17, 5, excelPath, 5)

        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)

    def exportDualValuesSheet1(self, selectedDataset, selectedNetwork, selectedScenarior, excelPath):
        '''
        This method is used to get data making DualValues_table.
        :param selectedDataset: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:

            sql = 'SELECT AttributeName, SourceName, InstanceName,MasterNetworkName,ScenarioName,MethodName,dualvaluemeaningCV ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "DualValues" ON "DualValues"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'LEFT JOIN "CV_DualValueMeaning" ON "CV_DualValueMeaning"."Name"= "DualValues"."dualvaluemeaningCV" '\
                'WHERE "Attributes"."AttributeDataTypeCV"="DualValues" AND "Datasets"."DatasetAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                .format(selectedDataset, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.dualvaluemeaningCV])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "DualValues")
                self.write2excel(complete_result, '4_DualValues', 9, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Error occurred in reading Data Structure.\n' + e.message)

    def exportTextConrolledSheet1(self, selectedDataset, selectedNetwork, selectedScenarior, excelPath):
        '''
        This method is used to get data making DescriptorValues_table.
        :param selectedDataset: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT Attributes.AttributeName, ObjectType, SourceName, InstanceName,MasterNetworkName,' \
                  'ScenarioName,MethodName,descriptorvalueCV ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "DescriptorValues" ON "DescriptorValues"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'LEFT JOIN "CV_DescriptorValues" ON "CV_DescriptorValues"."Name"= "DescriptorValues"."descriptorvalueCV" '\
                'WHERE "Attributes"."AttributeDataTypeCV"="DescriptorValues" AND "Datasets"."DatasetAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                .format(selectedDataset, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.descriptorvalueCV])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "DescriptorValues")
                self.write2excel(complete_result, '4_DescriptorValues', 13, 10, excelPath)
        except Exception as  e:
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)

    def exportNumericValuesheet1(self, selectedDataset, selectedNetwork, selectedScenarior, excelPath):
        '''
        This method is used to get data making Parameter.
        :param selectedDataset: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT "Datasets"."DatasetName", ObjectType,AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                  'ScenarioName,MethodName, NumericValue ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID"'\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "NumericValues" ON "NumericValues"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'WHERE "Attributes"."AttributeDataTypeCV"="Parameter" AND "Datasets"."DatasetAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                .format(selectedDataset, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.NumericValue])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "Parameter")
                self.write2excel(complete_result, '4_Parameter', 10, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportElectronicFilesSheet1(self, selectedDataset, selectedNetwork, selectedScenarior, excelPath):
        '''
        This method is used to get data making ElectronicFiles.
        :param selectedDataset: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                  'ScenarioName,MethodName,FileName, ElectronicFileFormatCV, "File"."Description" ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "Files" ON "File"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'Left JOIN "CV_ElectronicFormat" ON "CV_ElectronicFormat"."Name"="File"."ElectronicFileFormatCV" '\
                'WHERE "Attributes"."AttributeDataTypeCV"="File" AND "Datasets"."DatasetAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                .format(selectedDataset, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.FileName, row.ElectronicFileFormatCV, excelPath, row.Description])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "ElectronicFiles")
                self.write2excel(complete_result, '4_ElectronicFiles', 14, 10, excelPath)
        except Exception as  e:
            pass
            # print e
            # raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportSeasonalSheet1(self, selectedDataset, selectedNetwork, selectedScenarior, excelPath):
        '''
        This method is used to get data making SeasonalParameter.
        :param selectedDataset: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT ObjectType, AttributeName, SourceName, InstanceName,MasterNetworkName,' \
                  'ScenarioName,MethodName,SeasonName, SeasonValue, SeasonNameCV ' \
                'FROM "Attributes" '\
                'Left JOIN "ObjectTypes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "SeasonalParameters" ON "SeasonalParameters"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'WHERE "AttributeDataTypeCV"="SeasonalParameter" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                .format(selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.SeasonName, row.SeasonNameCV, row.SeasonValue])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "SeasonalParameter")
                self.write2excel(complete_result, '4_SeasonaNumericValues', 11, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)
    def exportTimeSeriesSheet1(self, selectedDataset, selectedNetwork, selectedScenarior, excelPath):
        '''
        This method is used to get data making TimeSeries.
        :param selectedDataset: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT DatasetName ObjectType, AttributeName, SourceName, InstanceName,WaterOrCalendarYear,' \
                  'ScenarioName,MethodName,AggregationStatisticCV, AggregationInterval, IntervalTimeUnitCV,' \
                  'IsRegular, NoDataValue, "TimeSeries"."Description" ' \
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'LEFT JOIN "TimeSeries" ON "TimeSeries"."DataValuesMapperID" = "DataValuesMapper"."DataValuesMapperID" '\
                'WHERE AttributeName!="ObjectInstances"  AND AttributeDataTypeCV="TimeSeries" ' \
                   'AND "DatasetAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                .format(selectedDataset, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)
            # nameResult = list()
            complete_result = list()
            for row in result:
                # isExisting = False
                # for name in nameResult:
                #     if name == row.InstanceName:
                #         isExisting = True
                #         break
                # if not isExisting:
                #     nameResult.append(row.InstanceName)
                complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName,
                                        row.AttributeName, row.SourceName, row.MethodName,
                                        row.WaterOrCalendarYear, row.AggregationStatisticCV, row.AggregationInterval,
                                        row.IntervalTimeUnit, row.IsRegular, row.NoDataValue, row.Description])
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "TimeSeries")
                self.write2excel(complete_result, '4_TimeSeries', 15, 10, excelPath)
        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)

    def exportMultiSheet1(self, selectedDataset, selectedNetwork, selectedScenarior, excelPath):
        '''
        This method is used to get data making MultiVariableSeries.
        :param selectedDataset: selected Model name
        :param selectedNetwork: selected Master Network name
        :param selectedScenarior: selected scenario Name
        :param excelPath: full path of excel file to export data
        :return: None
        '''
        try:
            sql = 'SELECT "ObjectTypes"."ObjectType", "Attributes"."AttributeName", SourceName, InstanceName,' \
                  'ScenarioName,MethodName, "AttributesColumns"."AttributeName" AS "ColumName", "AttributesColumns"."UnitNameCV" AS "ColUnitName", "Value","ValueOrder" '\
                'FROM "Datasets" '\
                'Left JOIN "ObjectTypes" ON "ObjectTypes"."DatasetID"="Datasets"."DatasetID" '\
                'Left JOIN "Attributes" ON "Attributes"."ObjectTypeID"="ObjectTypes"."ObjectTypeID" '\
                'Left JOIN "Mapping" ON "Mapping"."AttributeID"= "Attributes"."AttributeID" '\
                'Left JOIN "DataValuesMapper" ON "DataValuesMapper"."DataValuesMapperID"="Mapping"."DataValuesMapperID" '\
                'Left JOIN "ScenarioMapping" ON "ScenarioMapping"."MappingID"="Mapping"."MappingID" '\
                'Left JOIN "Scenarios" ON "Scenarios"."ScenarioID"="ScenarioMapping"."ScenarioID" '\
                'Left JOIN "MasterNetworks" ON "MasterNetworks"."MasterNetworkID"="Scenarios"."MasterNetworkID" '\
                'Left JOIN "Methods" ON "Methods"."MethodID"="Mapping"."MethodID" '\
                'Left JOIN "Sources" ON "Sources"."SourceID"="Mapping"."SourceID" '\
                'Left JOIN "Instances" ON "Instances"."InstanceID"="Mapping"."InstanceID" '\
                'Left JOIN "MultiAttributeSeries" ON "MultiAttributeSeries"."DataValuesMapperID"="DataValuesMapper"."DataValuesMapperID" '\
                'Left JOIN "DataValuesMapper" As "DataValuesMapperColumn" ON "DataValuesMapperColumn"."DataValuesMapperID"="MultiAttributeSeries"."AttriNameID" '\
                'Left JOIN "Mapping" As "MappingColumns" ON "MappingColumns"."DataValuesMapperID"="DataValuesMapperColumn"."DataValuesMapperID" '\
                'Left JOIN "Attributes" AS "AttributesColumns" ON "AttributesColumns"."AttributeID"="MappingColumns"."AttributeID" '\
                'Left JOIN "MultiAttributeSeriesValues" ON "MultiAttributeSeriesValues"."MultiAttributeSeriesID"="MultiAttributeSeries"."MultiAttributeSeriesID" '\
                'WHERE Attributes.AttributeDataTypeCV="MultiAttributeSeries" '\
                'AND "DatasetAcronym" = "{}" AND "MasterNetworkName" = "{}" AND "ScenarioName" = "{}"'\
                'Order By ScenarioName, AttributeName,ValueOrder asc '\
                .format(selectedDataset, selectedNetwork, selectedScenarior)

            result = self.session.execute(sql)

            '''Down Table(MultiVariableSeries_table Table) write'''
            complete_result = list()
            strAtrributName = ''
            valueOrder = None
            columnName = ''
            tempColumn = {}
            sourceName = ''
            i = 0
            currentrow = 0
            setNumber = 0
            for row in result:
                if row.ColumName == None or row.ColumName == "":
                    continue
                if strAtrributName != row.AttributeName:
                    strAtrributName = row.AttributeName
                    tempColumn[row.AttributeName] = []
                    tempColumn[row.AttributeName].append(row.ColumName)
                    columnName = row.ColumName
                if sourceName != row.ScenarioName:
                    sourceName = row.ScenarioName
                    setNumber = i
                    currentrow = 0
                if columnName != row.ColumName:
                    columnName = row.ColumName
                    currentrow = 0

                if row.ColumName in tempColumn[row.AttributeName]:
                    index = tempColumn[row.AttributeName].index(row.ColumName)
                    if index == 0:
                        complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName, row.AttributeName, row.SourceName,
                                        row.MethodName, row.Value])
                        i += 1
                    else:
                        complete_result[setNumber + currentrow].append(row.Value)
                        currentrow += 1
                else:
                    currentrow = 0
                    tempColumn[row.AttributeName].append(row.ColumName)
                    index = tempColumn[row.AttributeName].index(row.ColumName)
                    if index == 0:
                        complete_result.append([row.ObjectType, row.InstanceName, row.ScenarioName, row.AttributeName, row.SourceName,
                                        row.MethodName, row.Value])
                        i += 1
                    else:
                        complete_result[setNumber + currentrow].append(row.Value)
                        currentrow += 1
            if complete_result.__len__() > 0:
                self.isMatching_query(complete_result, "MultiVariableSeries")
                self.write2excel(complete_result, '4_MultiVariableSeries', 17, 15, excelPath)

            '''Up Table(AttributeName_column Table) write'''
            column_result = list()
            for key, columnItems in tempColumn.items():
                temList = list()
                temList.append(key)
                for item in columnItems:
                    temList.append(item)
                column_result.append(temList)
            if complete_result.__len__() > 0:
                self.write2excel(column_result, '4_MultiAttributeSeries', 17, 5, excelPath, 5)

        except Exception as  e:
            print e
            raise Exception('Erro occure in reading Data Structure.\n' + e.message)


    def write2excel(self, complete_result, xlsx_sheet_name, xls_sheet_num, startRowNum, excelPath, columnNum = 0):
        '''
        This method is used to write data specific sheet.
        :param complete_result: data to write within sheet
        :param xlsx_sheet_name: sheet name
        :param xls_sheet_num: sheet number to write
        :param startRowNum: start roe number to write
        :param excelPath: full path of excel file to export data
        :param columnNum: column number
        :return: None
        '''
        try:
            if excelPath.split('.')[-1] == 'xls':
                wb = open_workbook(excelPath)
                # workbook = copy(wb)
                workbook = wb
                try:
                    sheet = workbook.get_sheet(xls_sheet_num)
                except:
                    raise Exception("Please select a valid Excel File")
                for row_id, row in enumerate(complete_result):
                    for col_id, cell in enumerate(row):
                        sheet.write(row_id + startRowNum - 1, col_id + columnNum, cell)
                workbook.save(excelPath)
            else:
                book2 = load_workbook(excelPath)
                try:
                    sheet = book2.get_sheet_by_name(xlsx_sheet_name)
                except:
                    raise Exception('Output Sheet {} not found in Excel File \n Please select valid Excel File'.format(xlsx_sheet_name))

                for row_id, row in enumerate(complete_result):
                    for col_id, cell in enumerate(row):
                        sheet.cell(row=row_id + startRowNum, column=col_id + columnNum + 1, value=cell)
                book2.save(excelPath)

        except Exception as e:
            print e
            raise Exception(e.message)

    def isMatching_query(self, resultList, stateQuery):
        if (len(resultList) < 1):
            raise Exception("Nothing found in the {} selected WAMDAM database to match your request".format(stateQuery))

